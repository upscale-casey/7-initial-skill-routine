#!/usr/bin/env python3
"""
IR & External Relations Audit — Excel Workbook Builder

Generates a branded Upscale Excel workbook (.xlsx) from IR audit JSON data.
Uses Upscale brand colours and Lexend font throughout.

Upscale Brand Colours:
  - Eerie Black: #191919
  - Lime Green: #34C52A
  - Dark Cyan: #429792
  - Ivory: #FEFFEA
"""

import json
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# Brand Colours
EERIE_BLACK = "191919"
LIME_GREEN = "34C52A"
DARK_CYAN = "429792"
IVORY = "FEFFEA"
LIGHT_GRAY = "F5F5F5"

# Font
FONT_NAME = "Lexend"


def get_score_colour(score):
    """Return colour code based on score (1-10)."""
    if score >= 9:
        return LIME_GREEN
    elif score >= 7:
        return DARK_CYAN
    elif score >= 5:
        return "FFC7CE"  # Light red
    else:
        return "FF0000"  # Red


def create_header_style():
    """Create header cell style."""
    return Font(
        name=FONT_NAME,
        size=12,
        bold=True,
        color="FFFFFF"
    ), PatternFill(
        start_color=EERIE_BLACK,
        end_color=EERIE_BLACK,
        fill_type="solid"
    )


def create_subheader_style():
    """Create subheader cell style."""
    return Font(
        name=FONT_NAME,
        size=11,
        bold=True,
        color="FFFFFF"
    ), PatternFill(
        start_color=DARK_CYAN,
        end_color=DARK_CYAN,
        fill_type="solid"
    )


def create_section_title_style():
    """Create section title style."""
    return Font(
        name=FONT_NAME,
        size=14,
        bold=True,
        color=EERIE_BLACK
    ), None


def create_score_box_style(score):
    """Create style for score box."""
    colour = get_score_colour(score)
    return Font(
        name=FONT_NAME,
        size=28,
        bold=True,
        color="FFFFFF"
    ), PatternFill(
        start_color=colour,
        end_color=colour,
        fill_type="solid"
    )


def create_label_style():
    """Create label style."""
    return Font(
        name=FONT_NAME,
        size=10,
        bold=True,
        color=EERIE_BLACK
    ), None


def create_normal_style():
    """Create normal text style."""
    return Font(
        name=FONT_NAME,
        size=10,
        color=EERIE_BLACK
    ), None


def create_accent_fill():
    """Create accent background fill."""
    return PatternFill(
        start_color=LIGHT_GRAY,
        end_color=LIGHT_GRAY,
        fill_type="solid"
    )


def create_table_border():
    """Create table border style."""
    thin_border = Side(style='thin', color=EERIE_BLACK)
    return Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border
    )


def build_overview_sheet(wb, data):
    """Build Overview sheet."""
    ws = wb.active
    ws.title = "Overview"

    # Title
    ws['A1'] = f"{data['company_name']} — IR & External Relations Audit"
    ws['A1'].font = Font(name=FONT_NAME, size=16, bold=True, color=EERIE_BLACK)
    ws.merge_cells('A1:D1')

    # Metadata
    ws['A3'] = "Analysis Date:"
    ws['B3'] = data['analysis_date']
    ws['A4'] = "Industry:"
    ws['B4'] = data['industry']
    ws['A5'] = "IR Maturity:"
    ws['B5'] = data['ir_maturity'].replace('_', ' ').title()

    for row in range(3, 6):
        ws[f'A{row}'].font = create_label_style()[0]
        ws[f'B{row}'].font = create_normal_style()[0]

    # Overall Score
    ws['A7'] = "Overall Score"
    ws['A7'].font = create_section_title_style()[0]

    ws['A8'] = data['overall_score']
    ws['A8'].font = create_score_box_style(data['overall_score'])[0]
    ws['A8'].fill = create_score_box_style(data['overall_score'])[1]
    ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[8].height = 50

    # Section Scores
    ws['C7'] = "Section Scores"
    ws['C7'].font = create_section_title_style()[0]

    sections = [
        ('Governance', data['governance']['score']),
        ('Investor Base', data['investor_base']['score']),
        ('Valuation & Comparables', data['valuation_comparables']['score']),
        ('Commitments vs. Performance', data['commitments_performance']['score']),
        ('External Reputation', data['external_reputation']['score']),
    ]

    row = 8
    for section_name, score in sections:
        ws[f'C{row}'] = section_name
        ws[f'C{row}'].font = create_label_style()[0]
        ws[f'D{row}'] = score
        ws[f'D{row}'].font = create_normal_style()[0]
        ws[f'D{row}'].fill = PatternFill(
            start_color=get_score_colour(score),
            end_color=get_score_colour(score),
            fill_type="solid"
        )
        ws[f'D{row}'].font = Font(
            name=FONT_NAME, size=10, bold=True, color="FFFFFF"
        )
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    # Executive Summary
    ws['A15'] = "Executive Summary"
    ws['A15'].font = create_section_title_style()[0]
    ws.merge_cells('A16:D20')
    ws['A16'] = data['executive_summary']
    ws['A16'].font = create_normal_style()[0]
    ws['A16'].alignment = Alignment(wrap_text=True, vertical='top')

    # Top Priorities
    ws['A22'] = "Top Priorities"
    ws['A22'].font = create_section_title_style()[0]

    priority_row = 23
    for priority_item in data.get('top_priorities', []):
        ws[f'A{priority_row}'] = f"Priority {priority_item['priority']}: {priority_item['section']}"
        ws[f'A{priority_row}'].font = create_label_style()[0]
        ws[f'A{priority_row + 1}'] = priority_item['action']
        ws[f'A{priority_row + 1}'].font = create_normal_style()[0]
        ws.merge_cells(f'A{priority_row + 1}:D{priority_row + 1}')
        priority_row += 3

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12


def build_governance_sheet(wb, governance):
    """Build Governance section sheet."""
    ws = wb.create_sheet("→ Governance")

    # Title and Score
    ws['A1'] = "Governance"
    ws['A1'].font = create_section_title_style()[0]
    ws['D1'] = governance['score']
    ws['D1'].font = create_score_box_style(governance['score'])[0]
    ws['D1'].fill = create_score_box_style(governance['score'])[1]

    # Summary
    row = 3
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    ws[f'A{row}'] = governance['summary']
    ws[f'A{row}'].font = create_normal_style()[0]
    ws.merge_cells(f'A{row}:D{row + 2}')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    row += 4

    # Board Composition
    ws[f'A{row}'] = "Board Composition"
    ws[f'A{row}'].font = create_label_style()[0]
    row += 1
    board = governance.get('board_composition', {})
    board_info = [
        ('Total Members', board.get('total_members', 'N/A')),
        ('Independent %', f"{board.get('independent_pct', 0)}%"),
        ('Diversity Score', board.get('diversity_score', 'N/A')),
        ('Avg Tenure (years)', board.get('avg_tenure_years', 'N/A')),
    ]
    for label, value in board_info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = create_label_style()[0]
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    # ESG Disclosure
    ws[f'A{row}'] = "ESG Disclosure"
    ws[f'A{row}'].font = create_label_style()[0]
    row += 1
    esg = governance.get('esg_disclosure', {})
    ws[f'A{row}'] = "Level"
    ws[f'A{row}'].font = create_label_style()[0]
    ws[f'B{row}'] = esg.get('level', 'N/A')
    ws[f'B{row}'].font = create_normal_style()[0]
    row += 1

    frameworks = esg.get('frameworks_adopted', [])
    if frameworks:
        ws[f'A{row}'] = "Frameworks"
        ws[f'A{row}'].font = create_label_style()[0]
        ws[f'B{row}'] = ', '.join(frameworks)
        ws[f'B{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    # IR Function
    ws[f'A{row}'] = "IR Function"
    ws[f'A{row}'].font = create_label_style()[0]
    row += 1
    ir_func = governance.get('ir_function', {})
    ir_info = [
        ('Reports To', ir_func.get('reports_to', 'N/A')),
        ('Team Size', ir_func.get('team_size', 'N/A')),
        ('Maturity', ir_func.get('maturity', 'N/A')),
    ]
    for label, value in ir_info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = create_label_style()[0]
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = create_normal_style()[0]
        row += 1

    row += 2

    # Strengths
    ws[f'A{row}'] = "Strengths"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for strength in governance.get('strengths', []):
        ws[f'A{row}'] = f"• {strength.get('point', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    # Weaknesses
    ws[f'A{row}'] = "Weaknesses"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for weakness in governance.get('weaknesses', []):
        ws[f'A{row}'] = f"• {weakness.get('point', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    # Recommendations
    ws[f'A{row}'] = "Recommendations"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for rec in governance.get('recommendations', []):
        ws[f'A{row}'] = f"• {rec.get('action', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        ws[f'B{row}'] = f"Priority: {rec.get('priority', 'N/A')}"
        ws[f'B{row}'].font = Font(name=FONT_NAME, size=9, italic=True)
        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25


def build_investor_base_sheet(wb, investor_base):
    """Build Investor Base section sheet."""
    ws = wb.create_sheet("→ Investor Base")

    ws['A1'] = "Investor Base"
    ws['A1'].font = create_section_title_style()[0]
    ws['D1'] = investor_base['score']
    ws['D1'].font = create_score_box_style(investor_base['score'])[0]
    ws['D1'].fill = create_score_box_style(investor_base['score'])[1]

    row = 3
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    ws[f'A{row}'] = investor_base['summary']
    ws[f'A{row}'].font = create_normal_style()[0]
    ws.merge_cells(f'A{row}:D{row + 2}')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    row += 4

    # Ownership Structure
    ws[f'A{row}'] = "Ownership Structure"
    ws[f'A{row}'].font = create_label_style()[0]
    row += 1
    ownership = investor_base.get('ownership_structure', {})
    ownership_info = [
        ('Institutional %', f"{ownership.get('institutional_pct', 0)}%"),
        ('Retail %', f"{ownership.get('retail_pct', 0)}%"),
        ('Insider %', f"{ownership.get('insider_pct', 0)}%"),
        ('Top 10 Concentration %', f"{ownership.get('top_10_concentration_pct', 0)}%"),
    ]
    for label, value in ownership_info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = create_label_style()[0]
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    # Investor Sentiment
    ws[f'A{row}'] = "Investor Sentiment"
    ws[f'A{row}'].font = create_label_style()[0]
    ws[f'B{row}'] = investor_base.get('investor_sentiment', 'N/A')
    ws[f'B{row}'].font = create_normal_style()[0]
    row += 1

    # Activist Risk
    ws[f'A{row}'] = "Activist Risk"
    ws[f'A{row}'].font = create_label_style()[0]
    ws[f'B{row}'] = investor_base.get('activist_risk', 'N/A')
    ws[f'B{row}'].font = create_normal_style()[0]
    row += 2

    # Strengths, Weaknesses, Recommendations
    ws[f'A{row}'] = "Strengths"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for strength in investor_base.get('strengths', []):
        ws[f'A{row}'] = f"• {strength.get('point', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    ws[f'A{row}'] = "Weaknesses"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for weakness in investor_base.get('weaknesses', []):
        ws[f'A{row}'] = f"• {weakness.get('point', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    ws[f'A{row}'] = "Recommendations"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for rec in investor_base.get('recommendations', []):
        ws[f'A{row}'] = f"• {rec.get('action', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25


def build_valuation_sheet(wb, valuation):
    """Build Valuation & Comparables section sheet with comparables table."""
    ws = wb.create_sheet("→ Valuation")

    ws['A1'] = "Valuation & Comparables"
    ws['A1'].font = create_section_title_style()[0]
    ws['D1'] = valuation['score']
    ws['D1'].font = create_score_box_style(valuation['score'])[0]
    ws['D1'].fill = create_score_box_style(valuation['score'])[1]

    row = 3
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    ws[f'A{row}'] = valuation['summary']
    ws[f'A{row}'].font = create_normal_style()[0]
    ws.merge_cells(f'A{row}:D{row + 2}')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    row += 4

    # Current Valuation
    ws[f'A{row}'] = "Current Valuation"
    ws[f'A{row}'].font = create_label_style()[0]
    row += 1
    cur_val = valuation.get('current_valuation', {})
    cur_val_info = [
        ('Market Cap', f"${cur_val.get('market_cap', 'N/A'):,}" if cur_val.get('market_cap') is not None else "N/A"),
        ('Enterprise Value', f"${cur_val.get('enterprise_value', 'N/A'):,}" if cur_val.get('enterprise_value') is not None else "N/A"),
        ('Share Price', f"${cur_val.get('share_price', 'N/A')}" if cur_val.get('share_price') is not None else "N/A"),
    ]
    for label, value in cur_val_info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = create_label_style()[0]
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    # Key Multiples Table
    ws[f'A{row}'] = "Key Multiples"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    # Table Headers
    headers = ['Metric', 'Company', 'Peer Median', 'Peer Range', 'Assessment']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=DARK_CYAN, end_color=DARK_CYAN, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    # Multiples data
    for metric in valuation.get('key_multiples', []):
        ws[f'A{row}'] = metric.get('metric_name', '')
        ws[f'B{row}'] = metric.get('company_value', '')
        ws[f'C{row}'] = metric.get('peer_median', '')
        ws[f'D{row}'] = str(metric.get('peer_range', ''))
        ws[f'E{row}'] = metric.get('assessment', '')

        # Highlight assessment
        assessment = metric.get('assessment', '')
        if assessment == 'premium':
            colour = LIME_GREEN
        elif assessment == 'discount':
            colour = "FF0000"
        else:
            colour = LIGHT_GRAY
        ws[f'E{row}'].fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")

        for col in range(1, 6):
            ws.cell(row=row, column=col).font = create_normal_style()[0]
            ws.cell(row=row, column=col).border = create_table_border()
        row += 1

    row += 2

    # Strengths, Weaknesses, Recommendations
    ws[f'A{row}'] = "Strengths"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for strength in valuation.get('strengths', []):
        ws[f'A{row}'] = f"• {strength.get('point', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    ws[f'A{row}'] = "Weaknesses"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for weakness in valuation.get('weaknesses', []):
        ws[f'A{row}'] = f"• {weakness.get('point', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    ws[f'A{row}'] = "Recommendations"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for rec in valuation.get('recommendations', []):
        ws[f'A{row}'] = f"• {rec.get('action', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15


def build_commitments_sheet(wb, commitments):
    """Build Commitments vs. Performance section sheet with guidance tracking table."""
    ws = wb.create_sheet("→ Commitments")

    ws['A1'] = "Commitments vs. Performance"
    ws['A1'].font = create_section_title_style()[0]
    ws['D1'] = commitments['score']
    ws['D1'].font = create_score_box_style(commitments['score'])[0]
    ws['D1'].fill = create_score_box_style(commitments['score'])[1]

    row = 3
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    ws[f'A{row}'] = commitments['summary']
    ws[f'A{row}'].font = create_normal_style()[0]
    ws.merge_cells(f'A{row}:E{row + 2}')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    row += 4

    # Guidance Accuracy
    ws[f'A{row}'] = "Guidance Accuracy"
    ws[f'A{row}'].font = create_label_style()[0]
    accuracy = commitments.get('guidance_accuracy_pct', 0)
    ws[f'B{row}'] = f"{accuracy}%"
    ws[f'B{row}'].font = create_normal_style()[0]
    ws[f'B{row}'].fill = PatternFill(
        start_color=get_score_colour(accuracy / 10),
        end_color=get_score_colour(accuracy / 10),
        fill_type="solid"
    )
    ws[f'B{row}'].font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    row += 2

    # Guidance History Table
    ws[f'A{row}'] = "Guidance History"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    # Table Headers
    headers = ['Period', 'Metric', 'Guidance', 'Actual', 'Assessment']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=DARK_CYAN, end_color=DARK_CYAN, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    # Guidance data
    for guidance in commitments.get('guidance_history', []):
        ws[f'A{row}'] = guidance.get('period', '')
        ws[f'B{row}'] = guidance.get('metric', '')
        ws[f'C{row}'] = guidance.get('guidance', '')
        ws[f'D{row}'] = guidance.get('actual', '')
        ws[f'E{row}'] = guidance.get('assessment', '')

        # Colour-code assessment
        assessment = guidance.get('assessment', '')
        if assessment == 'beat':
            colour = LIME_GREEN
        elif assessment == 'met':
            colour = DARK_CYAN
        else:  # missed
            colour = "FF0000"
        ws[f'E{row}'].fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        ws[f'E{row}'].font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")

        for col in range(1, 6):
            ws.cell(row=row, column=col).font = create_normal_style()[0]
            ws.cell(row=row, column=col).border = create_table_border()
        row += 1

    row += 2

    # Credibility Assessment
    ws[f'A{row}'] = "Credibility Assessment"
    ws[f'A{row}'].font = create_label_style()[0]
    ws[f'B{row}'] = commitments.get('credibility_assessment', 'N/A')
    ws[f'B{row}'].font = create_normal_style()[0]
    row += 2

    # Strengths, Weaknesses, Recommendations
    ws[f'A{row}'] = "Strengths"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for strength in commitments.get('strengths', []):
        ws[f'A{row}'] = f"• {strength.get('point', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    ws[f'A{row}'] = "Weaknesses"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for weakness in commitments.get('weaknesses', []):
        ws[f'A{row}'] = f"• {weakness.get('point', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    ws[f'A{row}'] = "Recommendations"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for rec in commitments.get('recommendations', []):
        ws[f'A{row}'] = f"• {rec.get('action', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15


def build_reputation_sheet(wb, reputation):
    """Build External Reputation section sheet."""
    ws = wb.create_sheet("→ Reputation")

    ws['A1'] = "External Reputation"
    ws['A1'].font = create_section_title_style()[0]
    ws['D1'] = reputation['score']
    ws['D1'].font = create_score_box_style(reputation['score'])[0]
    ws['D1'].fill = create_score_box_style(reputation['score'])[1]

    row = 3
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    ws[f'A{row}'] = reputation['summary']
    ws[f'A{row}'].font = create_normal_style()[0]
    ws.merge_cells(f'A{row}:D{row + 2}')
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    row += 4

    # Media Sentiment
    ws[f'A{row}'] = "Media Sentiment"
    ws[f'A{row}'].font = create_label_style()[0]
    row += 1
    media = reputation.get('media_sentiment', {})
    media_info = [
        ('Overall', media.get('overall', 'N/A')),
        ('Volume Trend', media.get('volume_trend', 'N/A')),
    ]
    for label, value in media_info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = create_label_style()[0]
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    # Analyst Coverage
    ws[f'A{row}'] = "Analyst Coverage"
    ws[f'A{row}'].font = create_label_style()[0]
    row += 1
    analyst = reputation.get('analyst_coverage', {})
    analyst_info = [
        ('Total Analysts', analyst.get('total_analysts', 'N/A')),
        ('Buy %', f"{analyst.get('buy_pct', 0)}%"),
        ('Hold %', f"{analyst.get('hold_pct', 0)}%"),
        ('Sell %', f"{analyst.get('sell_pct', 0)}%"),
        ('Consensus', analyst.get('consensus', 'N/A')),
    ]
    for label, value in analyst_info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = create_label_style()[0]
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    # Stakeholder Trust
    ws[f'A{row}'] = "Stakeholder Trust"
    ws[f'A{row}'].font = create_label_style()[0]
    ws[f'B{row}'] = reputation.get('stakeholder_trust_assessment', 'N/A')
    ws[f'B{row}'].font = create_normal_style()[0]
    row += 2

    # Strengths, Weaknesses, Recommendations
    ws[f'A{row}'] = "Strengths"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for strength in reputation.get('strengths', []):
        ws[f'A{row}'] = f"• {strength.get('point', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    ws[f'A{row}'] = "Weaknesses"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for weakness in reputation.get('weaknesses', []):
        ws[f'A{row}'] = f"• {weakness.get('point', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    row += 1

    ws[f'A{row}'] = "Recommendations"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1
    for rec in reputation.get('recommendations', []):
        ws[f'A{row}'] = f"• {rec.get('action', '')}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25


def build_dynamics_sheet(wb, data):
    """Build Cross-Section Dynamics sheet."""
    ws = wb.create_sheet("Dynamics")

    ws['A1'] = "Cross-Section Dynamics"
    ws['A1'].font = create_section_title_style()[0]

    row = 3
    ws[f'A{row}'] = "Key Interactions & Interdependencies"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1

    for dynamic in data.get('cross_section_dynamics', []):
        ws[f'A{row}'] = f"• {dynamic}"
        ws[f'A{row}'].font = create_normal_style()[0]
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

    row += 2

    ws[f'A{row}'] = "Data Sources"
    ws[f'A{row}'].font = create_subheader_style()[0]
    ws[f'A{row}'].fill = create_subheader_style()[1]
    row += 1

    for source in data.get('data_sources', []):
        ws[f'A{row}'] = f"• {source}"
        ws[f'A{row}'].font = create_normal_style()[0]
        row += 1

    ws.column_dimensions['A'].width = 80


def build_excel(data, output_path):
    """Build complete Excel workbook from audit data."""
    wb = Workbook()

    # Overview sheet (first sheet)
    build_overview_sheet(wb, data)

    # Section sheets
    build_governance_sheet(wb, data['governance'])
    build_investor_base_sheet(wb, data['investor_base'])
    build_valuation_sheet(wb, data['valuation_comparables'])
    build_commitments_sheet(wb, data['commitments_performance'])
    build_reputation_sheet(wb, data['external_reputation'])

    # Dynamics sheet
    build_dynamics_sheet(wb, data)

    # Save
    wb.save(output_path)
    print(f"Excel workbook created: {output_path}")


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage: python build_excel.py <input_json_file> [output_xlsx_file]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else input_file.replace('.json', '.xlsx')

    try:
        with open(input_file, 'r') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"Error: File not found: {input_file}")
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON in {input_file}")
        sys.exit(1)

    build_excel(data, output_file)
    print(f"✓ Excel workbook completed: {output_file}")


if __name__ == '__main__':
    main()
