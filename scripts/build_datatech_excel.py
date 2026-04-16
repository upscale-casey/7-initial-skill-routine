#!/usr/bin/env python3
"""
Data & Technology Audit Excel Workbook Builder

Converts JSON audit data into a branded Upscale Excel workbook (.xlsx)
with consistent formatting, colors, and layout across all sheets.

Upscale Brand Colors:
  - Eerie Black: #191919
  - Lime Green: #34C52A
  - Dark Cyan: #429792
  - Ivory: #FEFFEA
Font: Lexend
"""

import json
import argparse
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, DEFAULT_FONT
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    exit(1)


class UpscaleBrandConfig:
    """Upscale brand color and style definitions."""

    COLORS = {
        'eerie_black': '191919',
        'lime_green': '34C52A',
        'dark_cyan': '429792',
        'ivory': 'FEFFEA',
        'light_gray': 'F5F5F5',
        'white': 'FFFFFF',
    }

    FONT_NAME = 'Lexend'

    @staticmethod
    def get_font(size: int = 11, bold: bool = False, color: str = '191919') -> Font:
        """Return a Lexend font with specified properties."""
        return Font(name=UpscaleBrandConfig.FONT_NAME, size=size, bold=bold, color=color)

    @staticmethod
    def get_fill(color_key: str) -> PatternFill:
        """Return a PatternFill with Upscale colors."""
        color = UpscaleBrandConfig.COLORS.get(color_key, '191919')
        return PatternFill(start_color=color, end_color=color, fill_type='solid')

    @staticmethod
    def get_border() -> Border:
        """Return a standard border."""
        thin_border = Side(style='thin', color='CCCCCC')
        return Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)


class DataTechAuditWorkbook:
    """Builds branded Excel workbook from audit JSON data."""

    def __init__(self, audit_data: Dict[str, Any]):
        self.data = audit_data
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        self.brand = UpscaleBrandConfig()

    def build(self) -> Workbook:
        """Build all sheets and return workbook."""
        self.create_overview_sheet()
        self.create_data_estate_sheet()
        self.create_martech_sheet()
        self.create_analytics_sheet()
        self.create_measurement_sheet()
        self.create_privacy_sheet()
        self.create_tech_debt_sheet()
        self.create_dynamics_sheet()
        return self.wb

    def _add_title(self, ws, title: str, row: int = 1):
        """Add a branded title row."""
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = self.brand.get_font(16, bold=True, color='FFFFFF')
        cell.fill = self.brand.get_fill('eerie_black')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 30

    def _add_section_header(self, ws, title: str, row: int):
        """Add a branded section header."""
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = self.brand.get_font(12, bold=True, color='FFFFFF')
        cell.fill = self.brand.get_fill('dark_cyan')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 22
        return row + 1

    def _add_score_badge(self, ws, score: int, label: str, col: str, row: int):
        """Add a score badge with color coding."""
        cell = ws[f'{col}{row}']
        cell.value = f'{label}: {score}/10'
        cell.font = self.brand.get_font(11, bold=True, color='FFFFFF')

        if score >= 8:
            cell.fill = self.brand.get_fill('lime_green')
        elif score >= 6:
            cell.fill = self.brand.get_fill('dark_cyan')
        else:
            cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')

        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.brand.get_border()
        ws.column_dimensions[col].width = 22

    def create_overview_sheet(self):
        """Create Overview sheet with scorecard and executive summary."""
        ws = self.wb.create_sheet('Overview', 0)
        ws.column_dimensions['A'].width = 3
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18

        row = 1
        self._add_title(ws, f"{self.data.get('company_name', 'Company')} - Data & Technology Audit", row)

        row = 3
        ws[f'A{row}'] = 'Analysis Date:'
        ws[f'B{row}'] = self.data.get('analysis_date', datetime.now().strftime('%Y-%m-%d'))
        ws[f'B{row}'].font = self.brand.get_font(11)

        row = 4
        ws[f'A{row}'] = 'Industry:'
        ws[f'B{row}'] = self.data.get('industry', 'N/A')
        ws[f'B{row}'].font = self.brand.get_font(11)

        row = 6
        row = self._add_section_header(ws, 'Scorecard', row)

        scores = [
            ('First-Party Data', self.data.get('first_party_data', {}).get('score', 0)),
            ('MarTech Stack', self.data.get('martech_stack', {}).get('score', 0)),
            ('Analytics Capability', self.data.get('analytics_capability', {}).get('score', 0)),
            ('Measurement & Attribution', self.data.get('measurement_attribution', {}).get('score', 0)),
            ('Privacy & Compliance', self.data.get('privacy_compliance', {}).get('score', 0)),
            ('Technology Debt', self.data.get('technology_debt', {}).get('score', 0)),
        ]

        cols = ['B', 'C', 'D', 'E', 'F', 'G']
        for i, (label, score) in enumerate(scores):
            self._add_score_badge(ws, score, label, cols[i], row)

        row += 2
        row = self._add_section_header(ws, 'Overall Assessment', row)

        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = f"Overall Score: {self.data.get('overall_score', 0)}/10"
        cell.font = self.brand.get_font(12, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = f"Digital Maturity: {self.data.get('digital_maturity', 'N/A').title()}"
        cell.font = self.brand.get_font(11)
        cell.alignment = Alignment(horizontal='left', vertical='top')

        row += 2
        row = self._add_section_header(ws, 'Executive Summary', row)

        ws.merge_cells(f'A{row}:G{row+3}')
        cell = ws[f'A{row}']
        cell.value = self.data.get('executive_summary', '')
        cell.font = self.brand.get_font(10)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 80

        row += 4
        row = self._add_section_header(ws, 'Top Priorities', row)

        for priority in self.data.get('top_priorities', []):
            ws.merge_cells(f'A{row}:G{row}')
            cell = ws[f'A{row}']
            cell.value = f"• {priority.get('priority', '')}"
            cell.font = self.brand.get_font(10)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.row_dimensions[row].height = 25
            row += 1

    def create_data_estate_sheet(self):
        """Create Data Estate sheet."""
        ws = self.wb.create_sheet('→ Data Estate')
        self._setup_sheet_layout(ws)

        row = 1
        self._add_title(ws, 'First-Party Data Estate', row)

        data = self.data.get('first_party_data', {})

        row = 3
        row = self._add_section_header(ws, f"Score: {data.get('score', 0)}/10", row)

        row = 1
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = 'First-Party Data Estate'
        cell.font = self.brand.get_font(16, bold=True, color='FFFFFF')
        cell.fill = self.brand.get_fill('eerie_black')
        ws.row_dimensions[row].height = 30

        row = 3
        ws[f'A{row}'] = f"Score: {data.get('score', 0)}/10"
        ws[f'A{row}'].font = self.brand.get_font(12, bold=True)

        row = 5
        row = self._add_section_header(ws, 'Summary', row)
        ws.merge_cells(f'A{row}:G{row+1}')
        cell = ws[f'A{row}']
        cell.value = data.get('summary', '')
        cell.font = self.brand.get_font(10)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 40

        row += 2
        self._add_strengths_weaknesses(ws, data, row)

    def create_martech_sheet(self):
        """Create MarTech Stack sheet with tools inventory table."""
        ws = self.wb.create_sheet('→ MarTech Stack')
        self._setup_sheet_layout(ws)

        row = 1
        self._add_title(ws, 'MarTech Stack Analysis', row)

        data = self.data.get('martech_stack', {})

        row = 3
        ws[f'A{row}'] = f"Total Tools: {data.get('total_tools', 0)} | Annual Spend: {data.get('annual_spend_estimate', 'N/A')}"
        ws[f'A{row}'].font = self.brand.get_font(11, bold=True)

        row = 4
        ws[f'A{row}'] = f"Score: {data.get('score', 0)}/10"
        ws[f'A{row}'].font = self.brand.get_font(11)

        row = 6
        row = self._add_section_header(ws, 'Tools Inventory', row)

        # Tools table headers
        headers = ['Tool', 'Category', 'Vendor', 'Utilization', 'Integration', 'Decision']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.brand.get_font(10, bold=True, color='FFFFFF')
            cell.fill = self.brand.get_fill('dark_cyan')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row += 1
        tools = data.get('tools', [])
        for tool in tools:
            ws.cell(row=row, column=1).value = tool.get('name', '')
            ws.cell(row=row, column=2).value = tool.get('category', '')
            ws.cell(row=row, column=3).value = tool.get('vendor', '')
            ws.cell(row=row, column=4).value = tool.get('utilization', '').title()
            ws.cell(row=row, column=5).value = tool.get('integration_quality', '').title()
            ws.cell(row=row, column=6).value = tool.get('decision', '').title()

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.font = self.brand.get_font(10)
                cell.border = self.brand.get_border()
                cell.alignment = Alignment(horizontal='left', vertical='center')

            row += 1

        row += 1
        row = self._add_section_header(ws, 'Analysis', row)
        self._add_strengths_weaknesses(ws, data, row)

    def create_analytics_sheet(self):
        """Create Analytics sheet with maturity model visualization."""
        ws = self.wb.create_sheet('→ Analytics')
        self._setup_sheet_layout(ws)

        row = 1
        self._add_title(ws, 'Analytics Capability & Maturity', row)

        data = self.data.get('analytics_capability', {})

        row = 3
        ws[f'A{row}'] = f"Maturity Level: {data.get('maturity_level', 'N/A').title()}"
        ws[f'A{row}'].font = self.brand.get_font(11, bold=True)

        row = 4
        ws[f'A{row}'] = f"Score: {data.get('score', 0)}/10"
        ws[f'A{row}'].font = self.brand.get_font(11)

        row = 6
        row = self._add_section_header(ws, 'Maturity Model', row)

        # Maturity stages
        stages = ['Descriptive', 'Diagnostic', 'Predictive', 'Prescriptive']
        stage_keys = ['descriptive', 'diagnostic', 'predictive', 'prescriptive']
        maturity_details = data.get('maturity_details', {})

        ws[f'A{row}'] = 'Stage'
        ws[f'B{row}'] = 'Capability'
        ws[f'C{row}'] = 'Tools'
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.font = self.brand.get_font(10, bold=True, color='FFFFFF')
            cell.fill = self.brand.get_fill('dark_cyan')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row += 1
        for stage, key in zip(stages, stage_keys):
            stage_data = maturity_details.get(key, {})
            ws[f'A{row}'] = stage
            ws[f'B{row}'] = stage_data.get('capability', 'N/A').title()
            ws[f'C{row}'] = ', '.join(stage_data.get('tools', []))

            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                cell.font = self.brand.get_font(10)
                cell.border = self.brand.get_border()
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            row += 1

        row += 1
        row = self._add_section_header(ws, 'Team & Organization', row)

        team = data.get('team_capability', {})
        ws[f'A{row}'] = 'Analysts:'
        ws[f'B{row}'] = team.get('analysts_count', 'N/A')
        ws[f'A{row+1}'] = 'Skill Level:'
        ws[f'B{row+1}'] = team.get('skill_level', 'N/A').title()
        ws[f'A{row+2}'] = 'Data Literacy:'
        ws[f'B{row+2}'] = data.get('data_literacy_org', 'N/A').title()

        for i in range(3):
            for col in ['A', 'B']:
                ws[f'{col}{row+i}'].font = self.brand.get_font(10)

        row += 3
        row = self._add_section_header(ws, 'Strengths & Weaknesses', row)
        self._add_strengths_weaknesses(ws, data, row)

    def create_measurement_sheet(self):
        """Create Measurement & Attribution sheet."""
        ws = self.wb.create_sheet('→ Measurement')
        self._setup_sheet_layout(ws)

        row = 1
        self._add_title(ws, 'Measurement & Attribution Framework', row)

        data = self.data.get('measurement_attribution', {})

        row = 3
        ws[f'A{row}'] = f"Attribution Model: {data.get('attribution_model', 'N/A').title()}"
        ws[f'A{row}'].font = self.brand.get_font(11, bold=True)

        row = 4
        ws[f'A{row}'] = f"Score: {data.get('score', 0)}/10"
        ws[f'A{row}'].font = self.brand.get_font(11)

        row = 6
        row = self._add_section_header(ws, 'Identity & Experimentation', row)

        identity = data.get('identity_resolution', {})
        ws[f'A{row}'] = 'Identity Resolution:'
        ws[f'B{row}'] = identity.get('capability', 'N/A').title()
        ws[f'A{row+1}'] = 'Method:'
        ws[f'B{row+1}'] = identity.get('method', 'N/A')

        row += 2
        experi = data.get('experimentation_capability', {})
        ws[f'A{row}'] = 'A/B Testing:'
        ws[f'B{row}'] = 'Yes' if experi.get('ab_testing') else 'No'
        ws[f'A{row+1}'] = 'Incrementality Testing:'
        ws[f'B{row+1}'] = 'Yes' if experi.get('incrementality_testing') else 'No'
        ws[f'A{row+2}'] = 'Marketing Mix Modeling:'
        ws[f'B{row+2}'] = 'Yes' if experi.get('mmm') else 'No'

        for i in range(5):
            for col in ['A', 'B']:
                ws[f'{col}{row-2+i}'].font = self.brand.get_font(10)

        row += 3
        row = self._add_section_header(ws, 'Strengths & Weaknesses', row)
        self._add_strengths_weaknesses(ws, data, row)

    def create_privacy_sheet(self):
        """Create Privacy & Compliance sheet."""
        ws = self.wb.create_sheet('→ Privacy')
        self._setup_sheet_layout(ws)

        row = 1
        self._add_title(ws, 'Privacy & Compliance', row)

        data = self.data.get('privacy_compliance', {})

        row = 3
        ws[f'A{row}'] = f"Regulatory Scope: {', '.join(data.get('regulatory_scope', []))}"
        ws[f'A{row}'].font = self.brand.get_font(11, bold=True)

        row = 4
        ws[f'A{row}'] = f"Score: {data.get('score', 0)}/10"
        ws[f'A{row}'].font = self.brand.get_font(11)

        row = 6
        row = self._add_section_header(ws, 'Compliance Status', row)

        cmp = data.get('consent_management', {})
        ws[f'A{row}'] = 'CMP Implemented:'
        ws[f'B{row}'] = 'Yes' if cmp.get('cmp_implemented') else 'No'
        ws[f'A{row+1}'] = 'CMP Vendor:'
        ws[f'B{row+1}'] = cmp.get('vendor', 'N/A')
        ws[f'A{row+2}'] = 'Privacy by Design:'
        ws[f'B{row+2}'] = 'Yes' if data.get('privacy_by_design') else 'No'
        ws[f'A{row+3}'] = 'Data Retention Policy:'
        ws[f'B{row+3}'] = data.get('data_retention_policy', 'N/A').title()

        for i in range(4):
            for col in ['A', 'B']:
                ws[f'{col}{row+i}'].font = self.brand.get_font(10)

        row += 4
        risk = data.get('risk_assessment', {})
        row = self._add_section_header(ws, f"Risk Level: {risk.get('overall_risk', 'N/A').title()}", row)

        row += 1
        row = self._add_section_header(ws, 'Strengths & Weaknesses', row)
        self._add_strengths_weaknesses(ws, data, row)

    def create_tech_debt_sheet(self):
        """Create Technology Debt sheet with legacy systems table."""
        ws = self.wb.create_sheet('→ Tech Debt')
        self._setup_sheet_layout(ws)

        row = 1
        self._add_title(ws, 'Technology Debt Assessment', row)

        data = self.data.get('technology_debt', {})
        debt_score = data.get('tech_debt_score', {})

        row = 3
        ws[f'A{row}'] = f"Tech Debt Score: {debt_score.get('value', 0)}/10 ({debt_score.get('percentile', 'N/A')})"
        ws[f'A{row}'].font = self.brand.get_font(11, bold=True)

        row = 4
        maint = data.get('maintenance_vs_innovation_ratio', {})
        ws[f'A{row}'] = f"Maintenance vs Innovation: {maint.get('maintenance_pct', 0)}% / {maint.get('innovation_pct', 0)}%"
        ws[f'A{row}'].font = self.brand.get_font(10)

        row = 6
        row = self._add_section_header(ws, 'Legacy Systems', row)

        headers = ['System', 'Age (Years)', 'Risk Level', 'Migration Plan']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.brand.get_font(10, bold=True, color='FFFFFF')
            cell.fill = self.brand.get_fill('dark_cyan')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row += 1
        for legacy in data.get('legacy_systems', []):
            ws.cell(row=row, column=1).value = legacy.get('system', '')
            ws.cell(row=row, column=2).value = legacy.get('age_years', '')
            ws.cell(row=row, column=3).value = legacy.get('risk_level', '').title()
            ws.cell(row=row, column=4).value = legacy.get('migration_plan', '').title()

            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = self.brand.get_font(10)
                cell.border = self.brand.get_border()

            row += 1

        row += 1
        row = self._add_section_header(ws, 'Code Quality', row)

        code = data.get('code_quality_indicators', {})
        ws[f'A{row}'] = 'Test Coverage:'
        ws[f'B{row}'] = code.get('test_coverage', 'N/A')
        ws[f'A{row+1}'] = 'Documentation:'
        ws[f'B{row+1}'] = code.get('documentation', 'N/A').title()
        ws[f'A{row+2}'] = 'Technical Standards:'
        ws[f'B{row+2}'] = 'Yes' if code.get('technical_standards') else 'No'

        for i in range(3):
            for col in ['A', 'B']:
                ws[f'{col}{row+i}'].font = self.brand.get_font(10)

        row += 3
        row = self._add_section_header(ws, 'Strengths & Weaknesses', row)
        self._add_strengths_weaknesses(ws, data, row)

    def create_dynamics_sheet(self):
        """Create Cross-Section Dynamics sheet."""
        ws = self.wb.create_sheet('Dynamics')
        self._setup_sheet_layout(ws)

        row = 1
        self._add_title(ws, 'Cross-Section Dynamics', row)

        row = 3
        row = self._add_section_header(ws, 'Key Interactions', row)

        for interaction in self.data.get('cross_section_dynamics', []):
            ws.merge_cells(f'A{row}:G{row}')
            cell = ws[f'A{row}']
            cell.value = f"• {interaction.get('interaction', '')}"
            cell.font = self.brand.get_font(10)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.row_dimensions[row].height = 25
            row += 1

            ws.merge_cells(f'A{row}:G{row}')
            cell = ws[f'A{row}']
            cell.value = f"Implication: {interaction.get('implication', '')}"
            cell.font = self.brand.get_font(9)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.row_dimensions[row].height = 20
            row += 1

    def _setup_sheet_layout(self, ws):
        """Setup basic sheet layout and column widths."""
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18

    def _add_strengths_weaknesses(self, ws, section_data: Dict, row: int):
        """Add strengths, weaknesses, and recommendations section."""
        strengths = section_data.get('strengths', [])
        weaknesses = section_data.get('weaknesses', [])
        recommendations = section_data.get('recommendations', [])

        row = self._add_section_header(ws, 'Strengths', row)
        for item in strengths:
            ws.merge_cells(f'A{row}:G{row}')
            cell = ws[f'A{row}']
            cell.value = f"• {item.get('title', '')}: {item.get('description', '')}"
            cell.font = self.brand.get_font(9)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.row_dimensions[row].height = 25
            row += 1

        row += 1
        row = self._add_section_header(ws, 'Weaknesses', row)
        for item in weaknesses:
            ws.merge_cells(f'A{row}:G{row}')
            cell = ws[f'A{row}']
            cell.value = f"• {item.get('title', '')}: {item.get('description', '')}"
            cell.font = self.brand.get_font(9)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.row_dimensions[row].height = 25
            row += 1

        row += 1
        row = self._add_section_header(ws, 'Recommendations', row)
        for item in recommendations:
            ws.merge_cells(f'A{row}:G{row}')
            cell = ws[f'A{row}']
            cell.value = f"• {item.get('action', '')}"
            cell.font = self.brand.get_font(9)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.row_dimensions[row].height = 25
            row += 1


def main():
    """Main entry point for Excel builder."""
    parser = argparse.ArgumentParser(
        description='Build branded Excel workbook from Data & Technology Audit JSON'
    )
    parser.add_argument('input_json', help='Path to input JSON audit file')
    parser.add_argument('-o', '--output', help='Path to output Excel file (default: derived from input)')

    args = parser.parse_args()

    input_path = Path(args.input_json)
    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}")
        exit(1)

    with open(input_path, 'r') as f:
        audit_data = json.load(f)

    output_path = args.output or str(input_path.with_suffix('.xlsx'))

    print(f"Building Excel workbook from {input_path}...")
    builder = DataTechAuditWorkbook(audit_data)
    workbook = builder.build()

    workbook.save(output_path)
    print(f"Excel workbook saved to: {output_path}")


if __name__ == '__main__':
    main()
