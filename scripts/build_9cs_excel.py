#!/usr/bin/env python3
"""
build_excel.py — Upscale-branded 9Cs Excel workbook generator
Usage: python build_excel.py --json <path_to_9cs.json> --out <output.xlsx>
"""
import argparse
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Upscale Brand Palette ─────────────────────────────────────────────────────
EERIE_BLACK   = "191919"
LIME_GREEN    = "34C52A"
DARK_CYAN     = "429792"
IVORY         = "FEFFEA"
PURPLE        = "7F2A90"
WHITE         = "FFFFFF"
GREY_DARK     = "333333"
GREY_MID      = "666666"
GREY_LIGHT    = "888888"

SCORE_GOOD_BG = "EDFAEC"
SCORE_MID_BG  = "E6F4F3"
SCORE_BAD_BG  = "F5EAF8"


def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def lfont(bold=False, size=10, color=EERIE_BLACK, italic=False):
    return Font(name="Lexend", bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def score_palette(score):
    if score >= 7:   return LIME_GREEN,  SCORE_GOOD_BG
    if score >= 5:   return DARK_CYAN,   SCORE_MID_BG
    return PURPLE, SCORE_BAD_BG

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def title_row(ws, row, text, height=36):
    ws.row_dimensions[row].height = height
    ws.merge_cells(f"A{row}:H{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name="Lexend", bold=True, size=14, color=WHITE)
    c.fill = fill(EERIE_BLACK)
    c.alignment = align("left", "center")

def subtitle_row(ws, row, text, height=20):
    ws.row_dimensions[row].height = height
    ws.merge_cells(f"A{row}:H{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name="Lexend", size=9, color=WHITE, italic=True)
    c.fill = fill(LIME_GREEN)
    c.alignment = align("left", "center")

def section_header(ws, row, text, height=20):
    ws.row_dimensions[row].height = height
    ws.merge_cells(f"A{row}:H{row}")
    c = ws[f"A{row}"]
    c.value = text.upper()
    c.font = Font(name="Lexend", bold=True, size=9, color=LIME_GREEN)
    c.fill = fill(IVORY)
    c.alignment = align("left", "center")

def label_value(ws, row, label, value, height=18):
    ws.row_dimensions[row].height = height
    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(name="Lexend", bold=True, size=9, color=EERIE_BLACK)
    c.fill = fill(IVORY)
    c.alignment = align("left", "center")
    c.border = thin_border()
    ws.merge_cells(f"C{row}:H{row}")
    c = ws[f"C{row}"]
    c.value = value
    c.font = Font(name="Lexend", size=9, color=GREY_DARK)
    c.fill = fill(WHITE)
    c.alignment = align("left", "center", wrap=True)
    c.border = thin_border()

def col_header(ws, row, specs, height=16):
    """specs = list of (col_start, col_end, label)"""
    ws.row_dimensions[row].height = height
    for start, end, label in specs:
        if start != end:
            ws.merge_cells(f"{start}{row}:{end}{row}")
        c = ws[f"{start}{row}"]
        c.value = label
        c.font = Font(name="Lexend", bold=True, size=9, color=WHITE)
        c.fill = fill(EERIE_BLACK)
        c.alignment = align("center", "center")
        c.border = thin_border()


# ── Overview sheet ────────────────────────────────────────────────────────────
def build_overview(wb, d):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":4,"B":20,"C":14,"D":14,"E":48,"F":1,"G":1,"H":1})
    r = 1

    title_row(ws, r, f"  {d['company_name'].upper()} — 9Cs ANALYSIS")
    r += 1
    subtitle_row(ws, r, f"  {d['industry']}  →  Analysis Date: {d['analysis_date']}")
    r += 2

    section_header(ws, r, "Scorecard Summary")
    r += 1

    col_header(ws, r, [("A","A","C"),("B","B","CATEGORY"),("C","C","SCORE"),("D","D","RATING"),("E","H","HEADLINE FINDING")])
    r += 1

    c_sections = [
        ("Company",       d.get("company",       {}), "company"),
        ("Customers",     d.get("customers",     {}), "customers"),
        ("Consumers",     d.get("consumers",     {}), "consumers"),
        ("Category",      d.get("category",      {}), "category"),
        ("Competitors",   d.get("competitors",   {}), "competitors"),
        ("Collaborators", d.get("collaborators", {}), "collaborators"),
        ("Climate",       d.get("climate",       {}), "climate"),
        ("Culture",       d.get("culture",       {}), "culture"),
        ("Community",     d.get("community",     {}), "community"),
    ]

    for idx, (label, section, _) in enumerate(c_sections):
        score = section.get("score", 5)
        ftxt, fbg = score_palette(score)
        ws.row_dimensions[r].height = 22

        c = ws[f"A{r}"]
        c.value = label[0]
        c.font = Font(name="Lexend", bold=True, size=11, color=WHITE)
        c.fill = fill(LIME_GREEN)
        c.alignment = align("center","center")
        c.border = thin_border()

        c = ws[f"B{r}"]
        c.value = label
        c.font = Font(name="Lexend", bold=True, size=9, color=EERIE_BLACK)
        c.fill = fill(IVORY)
        c.alignment = align("left","center")
        c.border = thin_border()

        c = ws[f"C{r}"]
        c.value = f"{score}/10"
        c.font = Font(name="Lexend", bold=True, size=13, color=ftxt)
        c.fill = fill(fbg)
        c.alignment = align("center","center")
        c.border = thin_border()

        rating_map = {LIME_GREEN:"Strong", DARK_CYAN:"Adequate", PURPLE:"Weak"}
        c = ws[f"D{r}"]
        c.value = rating_map.get(ftxt, "—")
        c.font = Font(name="Lexend", bold=True, size=9, color=ftxt)
        c.fill = fill(fbg)
        c.alignment = align("center","center")
        c.border = thin_border()

        ws.merge_cells(f"E{r}:H{r}")
        c = ws[f"E{r}"]
        headline = section.get("summary","")[:160]
        c.value = headline
        c.font = Font(name="Lexend", size=9, color=GREY_DARK)
        c.fill = fill(WHITE)
        c.alignment = align("left","center",wrap=True)
        c.border = thin_border()
        r += 1

    # Overall row
    overall = d.get("overall_score", 5)
    otxt, obg = score_palette(overall)
    ws.row_dimensions[r].height = 26
    ws.merge_cells(f"A{r}:B{r}")
    c = ws[f"A{r}"]
    c.value = "OVERALL SCORE"
    c.font = Font(name="Lexend", bold=True, size=9, color=WHITE)
    c.fill = fill(EERIE_BLACK)
    c.alignment = align("center","center")
    c.border = thin_border()

    c = ws[f"C{r}"]
    c.value = f"{overall}/10"
    c.font = Font(name="Lexend", bold=True, size=14, color=otxt)
    c.fill = fill(obg)
    c.alignment = align("center","center")
    c.border = thin_border()

    c = ws[f"D{r}"]
    c.value = d.get("overall_assessment","").replace("_"," ").upper()
    c.font = Font(name="Lexend", bold=True, size=9, color=otxt)
    c.fill = fill(obg)
    c.alignment = align("center","center")
    c.border = thin_border()

    ws.merge_cells(f"E{r}:H{r}")
    c = ws[f"E{r}"]
    c.value = d.get("executive_summary","")[:200]
    c.font = Font(name="Lexend", size=9, color=GREY_DARK)
    c.fill = fill(obg)
    c.alignment = align("left","center",wrap=True)
    c.border = thin_border()
    r += 2

    # Executive summary block
    section_header(ws, r, "Executive Summary")
    r += 1
    ws.row_dimensions[r].height = 80
    ws.merge_cells(f"A{r}:H{r}")
    c = ws[f"A{r}"]
    c.value = d.get("executive_summary","")
    c.font = Font(name="Lexend", size=9, color=GREY_DARK, italic=True)
    c.fill = fill(IVORY)
    c.alignment = align("left","top",wrap=True)
    c.border = thin_border()
    r += 2

    # Top priorities
    section_header(ws, r, "Top Priorities")
    r += 1
    col_header(ws, r, [("A","B","PRIORITY"),("C","C","C CATEGORY"),("D","H","RECOMMENDED ACTION")])
    r += 1

    priorities = d.get("top_priorities", [])
    for idx, p in enumerate(priorities[:5]):
        ws.row_dimensions[r].height = 22
        row_bg = IVORY if idx % 2 == 0 else WHITE

        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value = f"{idx+1} — {p.get('expected_impact','').upper()}"
        c.font = Font(name="Lexend", bold=True, size=9, color=LIME_GREEN)
        c.fill = fill(EERIE_BLACK)
        c.alignment = align("center","center",wrap=True)
        c.border = thin_border()

        c = ws[f"C{r}"]
        c.value = p.get("c_category","").title()
        c.font = Font(name="Lexend", size=9, color=DARK_CYAN)
        c.fill = fill(row_bg)
        c.alignment = align("center","center")
        c.border = thin_border()

        ws.merge_cells(f"D{r}:H{r}")
        c = ws[f"D{r}"]
        c.value = f"→  {p.get('priority','')}"
        c.font = Font(name="Lexend", size=9, color=GREY_DARK)
        c.fill = fill(row_bg)
        c.alignment = align("left","center",wrap=True)
        c.border = thin_border()
        r += 1


# ── Per-C sheet ───────────────────────────────────────────────────────────────
def build_c_sheet(wb, sheet_name, section_data, c_label):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":4,"B":22,"C":14,"D":14,"E":14,"F":12,"G":12,"H":14})
    score = section_data.get("score", 5)
    ftxt, fbg = score_palette(score)
    r = 1

    title_row(ws, r, f"  {c_label.upper()}  —  9Cs ANALYSIS")
    r += 1
    rating_str = "Strong" if score >= 7 else "Adequate" if score >= 5 else "Needs Improvement"
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"A{r}:H{r}")
    c = ws[f"A{r}"]
    c.value = f"  SCORE: {score}/10  —  {rating_str}"
    c.font = Font(name="Lexend", bold=True, size=11, color=WHITE)
    c.fill = fill(LIME_GREEN)
    c.alignment = align("left", "center")
    r += 2

    # Summary
    section_header(ws, r, "Summary")
    r += 1
    ws.row_dimensions[r].height = 66
    ws.merge_cells(f"A{r}:H{r}")
    c = ws[f"A{r}"]
    c.value = section_data.get("summary","")
    c.font = Font(name="Lexend", size=9, color=GREY_DARK, italic=True)
    c.fill = fill(IVORY)
    c.alignment = align("left","top",wrap=True)
    c.border = thin_border()
    r += 2

    # Key characteristics — pull whatever scalar fields exist
    section_header(ws, r, "Key Characteristics")
    r += 1
    skip_fields = {"summary","strengths","weaknesses","competitive_position","recommendations","score",
                   "core_offerings","price_points","primary_channels","core_capabilities","customer_segments",
                   "key_partners","key_competitors","stakeholder_groups","industry_trends","regulatory_changes",
                   "technology_shifts","esg_factors","sustainability_trends","macro_risks","behavior_patterns",
                   "pain_points","channel_partners"}
    for key, val in section_data.items():
        if key in skip_fields or not isinstance(val, (str, int, float)):
            continue
        label_value(ws, r, key.replace("_"," ").title(), str(val))
        r += 1
    r += 1

    # Strengths
    section_header(ws, r, "✓  Strengths")
    r += 1
    col_header(ws, r, [("A","B","STRENGTH"),("C","G","EVIDENCE"),("H","H","IMPACT")])
    r += 1
    for item in section_data.get("strengths", []):
        ws.row_dimensions[r].height = 22
        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value = item.get("point","")
        c.font = Font(name="Lexend", bold=True, size=9, color=EERIE_BLACK)
        c.fill = fill(SCORE_GOOD_BG)
        c.alignment = align("left","center",wrap=True)
        c.border = thin_border()

        ws.merge_cells(f"C{r}:G{r}")
        c = ws[f"C{r}"]
        c.value = item.get("evidence","")
        c.font = Font(name="Lexend", size=9, color=GREY_DARK, italic=True)
        c.fill = fill(WHITE)
        c.alignment = align("left","center",wrap=True)
        c.border = thin_border()

        impact = item.get("impact","medium")
        clr = LIME_GREEN if impact=="high" else DARK_CYAN if impact=="medium" else GREY_LIGHT
        c = ws[f"H{r}"]
        c.value = impact.upper()
        c.font = Font(name="Lexend", bold=True, size=8, color=clr)
        c.fill = fill(SCORE_GOOD_BG)
        c.alignment = align("center","center")
        c.border = thin_border()
        r += 1
    r += 1

    # Weaknesses
    section_header(ws, r, "✗  Weaknesses")
    r += 1
    col_header(ws, r, [("A","B","WEAKNESS"),("C","G","EVIDENCE"),("H","H","IMPACT")])
    r += 1
    for item in section_data.get("weaknesses", []):
        ws.row_dimensions[r].height = 22
        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value = item.get("point","")
        c.font = Font(name="Lexend", bold=True, size=9, color=EERIE_BLACK)
        c.fill = fill(SCORE_BAD_BG)
        c.alignment = align("left","center",wrap=True)
        c.border = thin_border()

        ws.merge_cells(f"C{r}:G{r}")
        c = ws[f"C{r}"]
        c.value = item.get("evidence","")
        c.font = Font(name="Lexend", size=9, color=GREY_DARK, italic=True)
        c.fill = fill(WHITE)
        c.alignment = align("left","center",wrap=True)
        c.border = thin_border()

        impact = item.get("impact","medium")
        clr = PURPLE if impact=="high" else DARK_CYAN if impact=="medium" else GREY_LIGHT
        c = ws[f"H{r}"]
        c.value = impact.upper()
        c.font = Font(name="Lexend", bold=True, size=8, color=clr)
        c.fill = fill(SCORE_BAD_BG)
        c.alignment = align("center","center")
        c.border = thin_border()
        r += 1
    r += 1

    # Competitive position (if present)
    comp_pos = section_data.get("competitive_position", [])
    if comp_pos:
        section_header(ws, r, "→  Competitive Position")
        r += 1
        col_header(ws, r, [("A","B","DIMENSION"),("C","E","KEY COMPETITORS"),("F","G","POSITION"),("H","H","NOTES")])
        r += 1
        pos_colors = {"leader":LIME_GREEN,"strong":LIME_GREEN,"average":DARK_CYAN,"weak":PURPLE,"lagging":PURPLE}
        for idx, item in enumerate(comp_pos):
            ws.row_dimensions[r].height = 22
            row_bg = IVORY if idx % 2 == 0 else WHITE
            ws.merge_cells(f"A{r}:B{r}")
            c = ws[f"A{r}"]
            c.value = item.get("dimension","")
            c.font = Font(name="Lexend", size=9, color=EERIE_BLACK)
            c.fill = fill(row_bg)
            c.alignment = align("left","center",wrap=True)
            c.border = thin_border()

            ws.merge_cells(f"C{r}:E{r}")
            c = ws[f"C{r}"]
            c.value = ", ".join(item.get("key_competitors",[]))
            c.font = Font(name="Lexend", size=9, color=GREY_MID)
            c.fill = fill(WHITE)
            c.alignment = align("left","center",wrap=True)
            c.border = thin_border()

            pos = item.get("position","average")
            ws.merge_cells(f"F{r}:G{r}")
            c = ws[f"F{r}"]
            c.value = pos.upper()
            c.font = Font(name="Lexend", bold=True, size=9, color=pos_colors.get(pos, GREY_MID))
            c.fill = fill(row_bg)
            c.alignment = align("center","center")
            c.border = thin_border()

            c = ws[f"H{r}"]
            c.value = item.get("notes","")
            c.font = Font(name="Lexend", size=8, color=DARK_CYAN, italic=True)
            c.fill = fill(WHITE)
            c.alignment = align("left","center",wrap=True)
            c.border = thin_border()
            r += 1
        r += 1

    # Recommendations
    section_header(ws, r, "→  Recommendations")
    r += 1
    col_header(ws, r, [("A","B","ACTION"),("C","F","RATIONALE"),("G","G","PRIORITY"),("H","H","TIMEFRAME")])
    r += 1
    pri_text = {"critical":PURPLE,"high":DARK_CYAN,"medium":GREY_MID,"low":GREY_LIGHT}
    pri_bg   = {"critical":SCORE_BAD_BG,"high":SCORE_MID_BG,"medium":IVORY,"low":WHITE}
    for item in section_data.get("recommendations", []):
        ws.row_dimensions[r].height = 30
        ws.merge_cells(f"A{r}:B{r}")
        c = ws[f"A{r}"]
        c.value = f"→  {item.get('action','')}"
        c.font = Font(name="Lexend", bold=True, size=9, color=EERIE_BLACK)
        c.fill = fill(IVORY)
        c.alignment = align("left","top",wrap=True)
        c.border = thin_border()

        ws.merge_cells(f"C{r}:F{r}")
        c = ws[f"C{r}"]
        c.value = item.get("rationale","")
        c.font = Font(name="Lexend", size=9, color=GREY_DARK, italic=True)
        c.fill = fill(WHITE)
        c.alignment = align("left","top",wrap=True)
        c.border = thin_border()

        pri = item.get("priority","medium")
        c = ws[f"G{r}"]
        c.value = pri.upper()
        c.font = Font(name="Lexend", bold=True, size=8, color=pri_text.get(pri, GREY_MID))
        c.fill = fill(pri_bg.get(pri, WHITE))
        c.alignment = align("center","center")
        c.border = thin_border()

        c = ws[f"H{r}"]
        c.value = item.get("timeframe","").replace("_"," ").title()
        c.font = Font(name="Lexend", size=8, color=DARK_CYAN)
        c.fill = fill(IVORY)
        c.alignment = align("center","center")
        c.border = thin_border()
        r += 1


# ── Cross-C Dynamics sheet ────────────────────────────────────────────────────
def build_cross_c_dynamics(wb, d):
    ws = wb.create_sheet("Cross-C Dynamics")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A":24,"B":52,"C":14})
    r = 1
    title_row(ws, r, "  CROSS-C SYNERGIES & CONFLICTS")
    r += 1
    subtitle_row(ws, r, "  Where the nine Cs reinforce or undermine each other")
    r += 2

    section_header(ws, r, "Analysis")
    r += 1
    col_header(ws, r, [("A","A","DYNAMIC / INTERACTION"),("B","B","DETAIL"),("C","C","TYPE")])
    r += 1

    for idx, syn in enumerate(d.get("cross_c_dynamics", [])):
        ws.row_dimensions[r].height = 56
        row_bg = IVORY if idx % 2 == 0 else WHITE

        # Heuristic: if it mentions conflict keywords, treat as conflict
        is_conflict = any(w in syn.lower() for w in ["conflict","undermin","squander","opaci","absent","dilut","lagging","despite"])
        status = "CONFLICT" if is_conflict else "SYNERGY"

        # Split on first sentence break for dynamic vs detail
        parts = syn.split(".", 1)
        dynamic = parts[0].strip()
        detail  = parts[1].strip() if len(parts) > 1 else ""

        c = ws[f"A{r}"]
        c.value = dynamic
        c.font = Font(name="Lexend", bold=True, size=9, color=EERIE_BLACK)
        c.fill = fill(SCORE_BAD_BG if is_conflict else SCORE_GOOD_BG)
        c.alignment = align("left","top",wrap=True)
        c.border = thin_border()

        c = ws[f"B{r}"]
        c.value = detail
        c.font = Font(name="Lexend", size=9, color=GREY_DARK, italic=True)
        c.fill = fill(WHITE)
        c.alignment = align("left","top",wrap=True)
        c.border = thin_border()

        clr = PURPLE if is_conflict else LIME_GREEN
        c = ws[f"C{r}"]
        c.value = status
        c.font = Font(name="Lexend", bold=True, size=10, color=clr)
        c.fill = fill(SCORE_BAD_BG if is_conflict else SCORE_GOOD_BG)
        c.alignment = align("center","center")
        c.border = thin_border()
        r += 1


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Build Upscale-branded 9Cs Excel from JSON")
    parser.add_argument("--json", required=True, help="Path to 9cs_analysis.json")
    parser.add_argument("--out",  required=True, help="Output .xlsx path")
    args = parser.parse_args()

    with open(args.json, "r", encoding="utf-8") as f:
        d = json.load(f)

    wb = Workbook()

    build_overview(wb, d)

    c_sheets = [
        ("① Company",       d.get("company",       {}), f"{d['company_name']} — Company"),
        ("② Customers",     d.get("customers",     {}), f"{d['company_name']} — Customers"),
        ("③ Consumers",     d.get("consumers",     {}), f"{d['company_name']} — Consumers"),
        ("④ Category",      d.get("category",      {}), f"{d['company_name']} — Category"),
        ("⑤ Competitors",   d.get("competitors",   {}), f"{d['company_name']} — Competitors"),
        ("⑥ Collaborators", d.get("collaborators", {}), f"{d['company_name']} — Collaborators"),
        ("⑦ Climate",       d.get("climate",       {}), f"{d['company_name']} — Climate"),
        ("⑧ Culture",       d.get("culture",       {}), f"{d['company_name']} — Culture"),
        ("⑨ Community",     d.get("community",     {}), f"{d['company_name']} — Community"),
    ]
    for sheet_name, section, label in c_sheets:
        build_c_sheet(wb, sheet_name, section, label)

    build_cross_c_dynamics(wb, d)

    wb.save(args.out)
    print(f"Saved: {args.out}")

if __name__ == "__main__":
    main()
