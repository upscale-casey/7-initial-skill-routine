#!/usr/bin/env python3
"""
Commercial Model Analysis - Excel Workbook Builder

Converts JSON analysis output into a branded Upscale Excel workbook with
professional formatting, color schemes, and layout.

Usage:
    python build_excel.py <input_json> <output_xlsx>
"""

import json
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Upscale Brand Colors
COLORS = {
    "eerie_black": "191919",
    "lime_green": "34C52A",
    "dark_cyan": "429792",
    "ivory": "FEFFEA",
    "light_gray": "F5F5F5",
    "dark_gray": "666666",
}

# Font settings
FONT_NAME = "Lexend"


def create_styles():
    """Create reusable style objects."""
    styles = {
        "title": Font(name=FONT_NAME, size=18, bold=True, color="FFFFFF"),
        "title_fill": PatternFill(start_color=COLORS["eerie_black"], end_color=COLORS["eerie_black"], fill_type="solid"),
        "heading": Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF"),
        "heading_fill": PatternFill(start_color=COLORS["dark_cyan"], end_color=COLORS["dark_cyan"], fill_type="solid"),
        "subheading": Font(name=FONT_NAME, size=11, bold=True, color=COLORS["eerie_black"]),
        "label": Font(name=FONT_NAME, size=10, bold=True, color=COLORS["eerie_black"]),
        "label_fill": PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid"),
        "normal": Font(name=FONT_NAME, size=10, color=COLORS["eerie_black"]),
        "score_high": Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF"),
        "score_high_fill": PatternFill(start_color=COLORS["lime_green"], end_color=COLORS["lime_green"], fill_type="solid"),
        "score_medium": Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF"),
        "score_medium_fill": PatternFill(start_color=COLORS["dark_cyan"], end_color=COLORS["dark_cyan"], fill_type="solid"),
        "accent_fill": PatternFill(start_color=COLORS["ivory"], end_color=COLORS["ivory"], fill_type="solid"),
        "border": Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        ),
    }
    return styles


def set_column_widths(ws, widths):
    """Set column widths."""
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def add_title_cell(ws, row, col, text, styles):
    """Add a title cell with styling."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = styles["title"]
    cell.fill = styles["title_fill"]
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def add_heading_cell(ws, row, col, text, styles, colspan=1):
    """Add a heading cell with styling."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = styles["heading"]
    cell.fill = styles["heading_fill"]
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if colspan > 1:
        for c in range(col + 1, col + colspan):
            ws.cell(row=row, column=c).fill = styles["heading_fill"]
    return cell


def add_label_value_pair(ws, row, col, label, value, styles):
    """Add a label-value pair with styling."""
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = styles["label"]
    label_cell.fill = styles["label_fill"]
    label_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    label_cell.border = styles["border"]

    value_cell = ws.cell(row=row, column=col + 1, value=value)
    value_cell.font = styles["normal"]
    value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    value_cell.border = styles["border"]


def create_overview_sheet(wb, data, styles):
    """Create the Overview sheet with dashboard."""
    ws = wb.create_sheet("Overview", 0)
    set_column_widths(ws, [25, 35, 12, 12, 12, 12, 12, 12])

    row = 1

    # Title
    ws.merge_cells(f"A{row}:H{row+1}")
    title = ws.cell(row=row, column=1)
    title.value = f"Commercial Model Analysis: {data.get('company_name', 'Company')}"
    title.font = styles["title"]
    title.fill = styles["title_fill"]
    title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 3

    # Executive Summary
    ws.merge_cells(f"A{row}:H{row}")
    add_heading_cell(ws, row, 1, "Executive Summary", styles, colspan=8)
    row += 1

    ws.merge_cells(f"A{row}:H{row+2}")
    summary = ws.cell(row=row, column=1)
    summary.value = data.get("executive_summary", "N/A")
    summary.font = styles["normal"]
    summary.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    summary.border = styles["border"]
    row += 4

    # Scorecard
    ws.merge_cells(f"A{row}:H{row}")
    add_heading_cell(ws, row, 1, "Section Scores", styles, colspan=8)
    row += 1

    # Score headers
    headers = ["Section", "Score", "Assessment"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = styles["heading"]
        cell.fill = styles["heading_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = styles["border"]
    row += 1

    # Scores
    sections = [
        ("Revenue Model", data.get("revenue_model", {}).get("score", "N/A")),
        ("Unit Economics", data.get("unit_economics", {}).get("score", "N/A")),
        ("Margin Profile", data.get("margin_profile", {}).get("score", "N/A")),
        ("Remuneration", data.get("remuneration", {}).get("score", "N/A")),
        ("Marketing ROI", data.get("marketing_roi", {}).get("score", "N/A")),
        ("Scenario Planning", data.get("scenario_planning", {}).get("score", "N/A")),
    ]

    for section, score in sections:
        ws.cell(row=row, column=1, value=section).font = styles["label"]
        ws.cell(row=row, column=1).border = styles["border"]

        score_cell = ws.cell(row=row, column=2, value=score)
        if isinstance(score, (int, float)):
            if score >= 7:
                score_cell.fill = styles["score_high_fill"]
                score_cell.font = styles["score_high"]
            else:
                score_cell.fill = styles["score_medium_fill"]
                score_cell.font = styles["score_medium"]
        score_cell.alignment = Alignment(horizontal="center", vertical="center")
        score_cell.border = styles["border"]

        assessment = (
            "Excellent"
            if isinstance(score, (int, float)) and score >= 8
            else "Strong"
            if isinstance(score, (int, float)) and score >= 6
            else "Adequate"
            if isinstance(score, (int, float)) and score >= 4
            else "Concerning"
        )
        ws.cell(row=row, column=3, value=assessment).font = styles["normal"]
        ws.cell(row=row, column=3).border = styles["border"]
        row += 1

    row += 1

    # Key Metrics
    ws.merge_cells(f"A{row}:H{row}")
    add_heading_cell(ws, row, 1, "Key Benchmarks", styles, colspan=8)
    row += 1

    benchmarks = data.get("key_benchmarks", [])
    if benchmarks:
        benchmark_headers = ["Metric", "Company Value", "Industry Benchmark", "Assessment"]
        for col, header in enumerate(benchmark_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = styles["heading"]
            cell.fill = styles["heading_fill"]
            cell.border = styles["border"]
        row += 1

        for benchmark in benchmarks[:10]:  # Show top 10
            ws.cell(row=row, column=1, value=benchmark.get("metric_name", "")).font = styles["normal"]
            ws.cell(row=row, column=1).border = styles["border"]
            ws.cell(row=row, column=2, value=benchmark.get("company_value", "")).font = styles["normal"]
            ws.cell(row=row, column=2).border = styles["border"]
            ws.cell(row=row, column=3, value=benchmark.get("industry_benchmark", "")).font = styles["normal"]
            ws.cell(row=row, column=3).border = styles["border"]
            ws.cell(row=row, column=4, value=benchmark.get("assessment", "")).font = styles["normal"]
            ws.cell(row=row, column=4).border = styles["border"]
            row += 1

    row += 1

    # Top Priorities
    ws.merge_cells(f"A{row}:H{row}")
    add_heading_cell(ws, row, 1, "Top Priorities", styles, colspan=8)
    row += 1

    priorities = data.get("top_priorities", [])
    if priorities:
        for priority in priorities[:5]:  # Show top 5
            ws.cell(row=row, column=1, value=priority.get("section", "")).font = styles["label"]
            ws.merge_cells(f"B{row}:H{row}")
            ws.cell(row=row, column=2, value=priority.get("expected_impact", "")).font = styles["normal"]
            row += 1


def create_section_sheet(wb, sheet_name, section_data, styles):
    """Create a section sheet with title, score, and details."""
    ws = wb.create_sheet(sheet_name)
    set_column_widths(ws, [25, 50])

    row = 1

    # Title with score
    ws.merge_cells(f"A{row}:B{row}")
    title_cell = ws.cell(row=row, column=1)
    score = section_data.get("score", "N/A")
    title_cell.value = f"{sheet_name} - Score: {score}/10"
    title_cell.font = styles["title"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 2

    # Summary
    ws.merge_cells(f"A{row}:B{row}")
    add_heading_cell(ws, row, 1, "Summary", styles, colspan=2)
    row += 1

    ws.merge_cells(f"A{row}:B{row+2}")
    summary_cell = ws.cell(row=row, column=1)
    summary_cell.value = section_data.get("summary", "N/A")
    summary_cell.font = styles["normal"]
    summary_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    summary_cell.border = styles["border"]
    row += 4

    # Key Metrics - varies by section
    if "key_metrics" in section_data:
        ws.merge_cells(f"A{row}:B{row}")
        add_heading_cell(ws, row, 1, "Key Metrics", styles, colspan=2)
        row += 1
        for key, value in section_data["key_metrics"].items():
            add_label_value_pair(ws, row, 1, key, str(value), styles)
            row += 1
        row += 1

    # Strengths
    if section_data.get("strengths"):
        ws.merge_cells(f"A{row}:B{row}")
        add_heading_cell(ws, row, 1, "Strengths", styles, colspan=2)
        row += 1
        for strength in section_data["strengths"]:
            add_label_value_pair(
                ws,
                row,
                1,
                strength.get("point", ""),
                f"{strength.get('evidence', '')} (Impact: {strength.get('impact', 'N/A')})",
                styles,
            )
            row += 1
        row += 1

    # Weaknesses
    if section_data.get("weaknesses"):
        ws.merge_cells(f"A{row}:B{row}")
        add_heading_cell(ws, row, 1, "Weaknesses", styles, colspan=2)
        row += 1
        for weakness in section_data["weaknesses"]:
            add_label_value_pair(
                ws,
                row,
                1,
                weakness.get("point", ""),
                f"{weakness.get('evidence', '')} (Impact: {weakness.get('impact', 'N/A')})",
                styles,
            )
            row += 1
        row += 1

    # Recommendations
    if section_data.get("recommendations"):
        ws.merge_cells(f"A{row}:B{row}")
        add_heading_cell(ws, row, 1, "Recommendations", styles, colspan=2)
        row += 1
        for rec in section_data["recommendations"]:
            add_label_value_pair(
                ws,
                row,
                1,
                rec.get("action", ""),
                f"Priority: {rec.get('priority', 'N/A')} | Effort: {rec.get('effort', 'N/A')} | Timeframe: {rec.get('timeframe', 'N/A')}",
                styles,
            )
            row += 1


def create_scenarios_sheet(wb, scenario_data, styles):
    """Create the Scenarios comparison sheet."""
    ws = wb.create_sheet("Scenarios")
    set_column_widths(ws, [25, 25, 25, 25])

    row = 1

    # Title
    ws.merge_cells(f"A{row}:D{row}")
    title_cell = ws.cell(row=row, column=1)
    title_cell.value = "Scenario Planning: Base / Bull / Bear"
    title_cell.font = styles["title"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 2

    # Scenario headers
    scenario_names = ["Base Case", "Bull Case", "Bear Case"]
    for col, name in enumerate(scenario_names, 1):
        cell = ws.cell(row=row, column=col + 1)
        cell.value = name
        cell.font = styles["heading"]
        cell.fill = styles["heading_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    # Assumptions
    ws.cell(row=row, column=1, value="Assumptions").font = styles["label"]
    ws.cell(row=row, column=1).fill = styles["label_fill"]

    cases = ["base_case", "bull_case", "bear_case"]
    for col, case_key in enumerate(cases, 2):
        case = scenario_data.get(case_key, {})
        assumptions = case.get("assumptions", [])
        assumptions_text = "\n".join(assumptions[:3]) if assumptions else "N/A"
        cell = ws.cell(row=row, column=col)
        cell.value = assumptions_text
        cell.font = styles["normal"]
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = styles["border"]
    row += 1

    # Revenue Forecast
    ws.cell(row=row, column=1, value="Revenue Forecast").font = styles["label"]
    ws.cell(row=row, column=1).fill = styles["label_fill"]

    for col, case_key in enumerate(cases, 2):
        case = scenario_data.get(case_key, {})
        revenue = case.get("revenue_forecast", "N/A")
        cell = ws.cell(row=row, column=col)
        cell.value = revenue
        cell.font = styles["normal"]
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = styles["border"]
    row += 1

    # Margin Forecast
    ws.cell(row=row, column=1, value="Margin Forecast").font = styles["label"]
    ws.cell(row=row, column=1).fill = styles["label_fill"]

    for col, case_key in enumerate(cases, 2):
        case = scenario_data.get(case_key, {})
        margin = case.get("margin_forecast", "N/A")
        cell = ws.cell(row=row, column=col)
        cell.value = margin
        cell.font = styles["normal"]
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = styles["border"]
    row += 2

    # Key Sensitivities
    ws.merge_cells(f"A{row}:D{row}")
    add_heading_cell(ws, row, 1, "Key Sensitivities", styles, colspan=4)
    row += 1

    sensitivities = scenario_data.get("key_sensitivities", [])
    if sensitivities:
        ws.cell(row=row, column=1, value="Variable").font = styles["heading"]
        ws.cell(row=row, column=1).fill = styles["heading_fill"]
        ws.cell(row=row, column=2, value="Impact on Revenue").font = styles["heading"]
        ws.cell(row=row, column=2).fill = styles["heading_fill"]
        ws.cell(row=row, column=3, value="Impact on Margin").font = styles["heading"]
        ws.cell(row=row, column=3).fill = styles["heading_fill"]
        row += 1

        for sensitivity in sensitivities:
            ws.cell(row=row, column=1, value=sensitivity.get("variable", "")).font = styles["normal"]
            ws.cell(row=row, column=2, value=sensitivity.get("impact_on_revenue", "")).font = styles["normal"]
            ws.cell(row=row, column=3, value=sensitivity.get("impact_on_margin", "")).font = styles["normal"]
            row += 1


def create_benchmarks_sheet(wb, benchmarks, styles):
    """Create the Benchmarks comparison sheet."""
    ws = wb.create_sheet("Benchmarks")
    set_column_widths(ws, [25, 25, 25, 20, 15])

    row = 1

    # Title
    ws.merge_cells(f"A{row}:E{row}")
    title_cell = ws.cell(row=row, column=1)
    title_cell.value = "Key Benchmarks & Industry Comparison"
    title_cell.font = styles["title"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 2

    # Headers
    headers = ["Metric", "Company Value", "Industry Benchmark", "Percentile", "Assessment"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = styles["heading"]
        cell.fill = styles["heading_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = styles["border"]
    row += 1

    # Benchmark rows
    for benchmark in benchmarks:
        ws.cell(row=row, column=1, value=benchmark.get("metric_name", "")).font = styles["normal"]
        ws.cell(row=row, column=1).border = styles["border"]
        ws.cell(row=row, column=2, value=benchmark.get("company_value", "")).font = styles["normal"]
        ws.cell(row=row, column=2).border = styles["border"]
        ws.cell(row=row, column=3, value=benchmark.get("industry_benchmark", "")).font = styles["normal"]
        ws.cell(row=row, column=3).border = styles["border"]
        ws.cell(row=row, column=4, value=benchmark.get("percentile_estimate", "")).font = styles["normal"]
        ws.cell(row=row, column=4).border = styles["border"]

        assessment_cell = ws.cell(row=row, column=5)
        assessment_cell.value = benchmark.get("assessment", "")
        assessment_cell.font = styles["normal"]
        assessment_cell.border = styles["border"]
        if benchmark.get("assessment") == "above_benchmark":
            assessment_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=COLORS["lime_green"])
        row += 1


def create_dynamics_sheet(wb, dynamics, styles):
    """Create the Cross-Section Dynamics sheet."""
    ws = wb.create_sheet("Dynamics")
    set_column_widths(ws, [80])

    row = 1

    # Title
    ws.merge_cells(f"A{row}:A{row+1}")
    title_cell = ws.cell(row=row, column=1)
    title_cell.value = "Cross-Section Dynamics & Relationships"
    title_cell.font = styles["title"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 3

    # Dynamics
    if dynamics:
        for i, dynamic in enumerate(dynamics, 1):
            ws.merge_cells(f"A{row}:A{row}")
            header_cell = ws.cell(row=row, column=1)
            header_cell.value = f"Dynamic #{i}"
            header_cell.font = styles["label"]
            header_cell.fill = styles["label_fill"]
            row += 1

            ws.merge_cells(f"A{row}:A{row+1}")
            detail_cell = ws.cell(row=row, column=1)
            detail_cell.value = dynamic
            detail_cell.font = styles["normal"]
            detail_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            detail_cell.border = styles["border"]
            row += 3


def build_workbook(json_file, output_file):
    """Build the complete Excel workbook from JSON data."""
    with open(json_file, "r") as f:
        data = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    styles = create_styles()

    # Create sheets in order
    create_overview_sheet(wb, data, styles)

    # Section sheets
    if data.get("revenue_model"):
        create_section_sheet(wb, "Revenue Model", data["revenue_model"], styles)

    if data.get("unit_economics"):
        create_section_sheet(wb, "Unit Economics", data["unit_economics"], styles)

    if data.get("margin_profile"):
        create_section_sheet(wb, "Margin Profile", data["margin_profile"], styles)

    if data.get("remuneration"):
        create_section_sheet(wb, "Remuneration", data["remuneration"], styles)

    if data.get("marketing_roi"):
        create_section_sheet(wb, "Marketing ROI", data["marketing_roi"], styles)

    if data.get("scenario_planning"):
        create_scenarios_sheet(wb, data["scenario_planning"], styles)

    # Reference sheets
    if data.get("key_benchmarks"):
        create_benchmarks_sheet(wb, data["key_benchmarks"], styles)

    if data.get("cross_section_dynamics"):
        create_dynamics_sheet(wb, data["cross_section_dynamics"], styles)

    # Save
    wb.save(output_file)
    print(f"Excel workbook created: {output_file}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python build_excel.py <input_json> <output_xlsx>")
        sys.exit(1)

    json_file = sys.argv[1]
    output_file = sys.argv[2]

    try:
        build_workbook(json_file, output_file)
    except Exception as e:
        print(f"Error building workbook: {e}", file=sys.stderr)
        sys.exit(1)
