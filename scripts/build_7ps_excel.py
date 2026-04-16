#!/usr/bin/env python3
"""
Marketing Audit (7Ps) Excel Workbook Builder

Generates a branded Excel workbook from a completed JSON audit object.
Uses Upscale brand palette: Eerie Black #191919, Lime Green #34C52A,
Dark Cyan #429792, Ivory #FEFFEA.
Font: Lexend throughout.
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# Upscale Brand Palette
COLOR_EERIE_BLACK = "FF191919"
COLOR_LIME_GREEN = "FF34C52A"
COLOR_DARK_CYAN = "FF429792"
COLOR_IVORY = "FFFEFFEA"
COLOR_WHITE = "FFFFFFFF"
COLOR_LIGHT_GRAY = "FFF5F5F5"

FONT_FAMILY = "Lexend"


def get_header_font():
    """Return header font (bold, white, Lexend)."""
    return Font(name=FONT_FAMILY, size=14, bold=True, color=COLOR_WHITE)


def get_subheader_font():
    """Return subheader font (bold, Lexend)."""
    return Font(name=FONT_FAMILY, size=11, bold=True, color=COLOR_EERIE_BLACK)


def get_body_font():
    """Return body font (Lexend)."""
    return Font(name=FONT_FAMILY, size=10, color=COLOR_EERIE_BLACK)


def get_header_fill():
    """Return header fill (Eerie Black)."""
    return PatternFill(start_color=COLOR_EERIE_BLACK, end_color=COLOR_EERIE_BLACK, fill_type="solid")


def get_subheader_fill():
    """Return subheader fill (Dark Cyan)."""
    return PatternFill(start_color=COLOR_DARK_CYAN, end_color=COLOR_DARK_CYAN, fill_type="solid")


def get_accent_fill():
    """Return accent fill (Lime Green)."""
    return PatternFill(start_color=COLOR_LIME_GREEN, end_color=COLOR_LIME_GREEN, fill_type="solid")


def get_light_fill():
    """Return light background fill (Ivory)."""
    return PatternFill(start_color=COLOR_IVORY, end_color=COLOR_IVORY, fill_type="solid")


def get_centered_alignment():
    """Return centered alignment."""
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def get_left_alignment():
    """Return left alignment with text wrap."""
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


def get_thin_border():
    """Return thin border."""
    thin = Side(style="thin", color="FFD3D3D3")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def score_to_color(score):
    """Map score (1-10) to a color gradient."""
    if score is None:
        return COLOR_LIGHT_GRAY
    if score >= 8:
        return COLOR_LIME_GREEN
    elif score >= 6:
        return COLOR_DARK_CYAN
    elif score >= 4:
        return COLOR_IVORY
    else:
        return COLOR_LIGHT_GRAY


def load_audit_json(json_path):
    """Load audit JSON from file."""
    with open(json_path, 'r') as f:
        return json.load(f)


def create_overview_sheet(wb, audit):
    """Create the Overview scorecard sheet."""
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_properties.tabColor = COLOR_LIME_GREEN

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20

    # Title
    ws['A1'] = f"Marketing Audit (7Ps) - {audit.get('company_name', 'Company')}"
    ws['A1'].font = get_header_font()
    ws['A1'].fill = get_header_fill()
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = get_centered_alignment()

    # Metadata
    ws['A2'] = f"Industry: {audit.get('industry', 'N/A')}"
    ws['A2'].font = get_body_font()
    ws['A3'] = f"Analysis Date: {audit.get('analysis_date', 'N/A')}"
    ws['A3'].font = get_body_font()

    # Executive Summary
    ws['A5'] = "Executive Summary"
    ws['A5'].font = get_subheader_font()
    ws['A5'].fill = get_subheader_fill()
    ws['A5'].font = Font(name=FONT_FAMILY, size=11, bold=True, color=COLOR_WHITE)
    ws.merge_cells('A5:C5')

    ws['A6'] = audit.get('executive_summary', '')
    ws['A6'].alignment = get_left_alignment()
    ws['A6'].font = get_body_font()
    ws.merge_cells('A6:C6')
    ws.row_dimensions[6].height = 60

    # Scorecard Header
    ws['A8'] = "Section"
    ws['B8'] = "Score"
    ws['C8'] = "Assessment"
    for col in ['A', 'B', 'C']:
        ws[f'{col}8'].font = get_header_font()
        ws[f'{col}8'].fill = get_header_fill()
        ws[f'{col}8'].alignment = get_centered_alignment()

    # Strategic
    row = 9
    ws[f'A{row}'] = "STRATEGIC & POSITIONING"
    ws[f'A{row}'].font = get_subheader_font()
    ws[f'A{row}'].fill = get_accent_fill()
    ws[f'A{row}'].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
    ws.merge_cells(f'A{row}:C{row}')
    row += 1

    sections_strategic = [
        ("Strategic Assessment", audit.get('strategic_assessment', {}).get('score')),
    ]

    for section_name, score in sections_strategic:
        ws[f'A{row}'] = section_name
        ws[f'B{row}'] = score if score else "—"
        ws[f'C{row}'] = audit.get('overall_assessment', 'N/A') if section_name == "Strategic Assessment" else ""
        ws[f'B{row}'].fill = PatternFill(start_color=score_to_color(score),
                                         end_color=score_to_color(score), fill_type="solid")
        ws[f'A{row}'].alignment = get_left_alignment()
        ws[f'B{row}'].alignment = get_centered_alignment()
        ws[f'C{row}'].alignment = get_left_alignment()
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = get_body_font()
            ws[f'{col}{row}'].border = get_thin_border()
        row += 1

    # 7Ps
    ws[f'A{row}'] = "7Ps MARKETING MIX"
    ws[f'A{row}'].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
    ws[f'A{row}'].fill = get_accent_fill()
    ws.merge_cells(f'A{row}:C{row}')
    row += 1

    sections_7ps = [
        ("Product", audit.get('product', {}).get('score')),
        ("Price", audit.get('price', {}).get('score')),
        ("Place", audit.get('place', {}).get('score')),
        ("Promotion", audit.get('promotion', {}).get('score')),
        ("People", audit.get('people', {}).get('score')),
        ("Process", audit.get('process', {}).get('score')),
        ("Physical Evidence", audit.get('physical_evidence', {}).get('score')),
    ]

    for section_name, score in sections_7ps:
        ws[f'A{row}'] = section_name
        ws[f'B{row}'] = score if score else "—"
        ws[f'C{row}'] = ""
        ws[f'B{row}'].fill = PatternFill(start_color=score_to_color(score),
                                         end_color=score_to_color(score), fill_type="solid")
        ws[f'A{row}'].alignment = get_left_alignment()
        ws[f'B{row}'].alignment = get_centered_alignment()
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = get_body_font()
            ws[f'{col}{row}'].border = get_thin_border()
        row += 1

    # Performance
    ws[f'A{row}'] = "PERFORMANCE & OPTIMIZATION"
    ws[f'A{row}'].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
    ws[f'A{row}'].fill = get_accent_fill()
    ws.merge_cells(f'A{row}:C{row}')
    row += 1

    sections_perf = [
        ("Brand Health", audit.get('brand_health', {}).get('score')),
        ("Customer Journey", audit.get('customer_journey', {}).get('score')),
        ("Budget Analysis", audit.get('budget_analysis', {}).get('score')),
        ("Performance Metrics", audit.get('performance_metrics', {}).get('score')),
    ]

    for section_name, score in sections_perf:
        ws[f'A{row}'] = section_name
        ws[f'B{row}'] = score if score else "—"
        ws[f'C{row}'] = ""
        ws[f'B{row}'].fill = PatternFill(start_color=score_to_color(score),
                                         end_color=score_to_color(score), fill_type="solid")
        ws[f'A{row}'].alignment = get_left_alignment()
        ws[f'B{row}'].alignment = get_centered_alignment()
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = get_body_font()
            ws[f'{col}{row}'].border = get_thin_border()
        row += 1

    # Overall
    row += 1
    ws[f'A{row}'] = "OVERALL SCORE"
    ws[f'B{row}'] = audit.get('overall_score', '—')
    ws[f'C{row}'] = audit.get('overall_assessment', '—')
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = get_header_font()
        ws[f'{col}{row}'].fill = get_header_fill()
        ws[f'{col}{row}'].alignment = get_centered_alignment()
        ws[f'{col}{row}'].border = get_thin_border()

    row += 2
    ws[f'A{row}'] = "Marketing Maturity:"
    ws[f'B{row}'] = audit.get('marketing_maturity', '—')
    ws[f'A{row}'].font = get_subheader_font()
    ws[f'B{row}'].font = get_body_font()


def create_detail_sheet(wb, sheet_name, section_data):
    """Create a detail sheet for a section."""
    ws = wb.create_sheet(sheet_name)

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50

    row = 1

    # Title
    ws[f'A{row}'] = sheet_name
    ws[f'A{row}'].font = get_header_font()
    ws[f'A{row}'].fill = get_header_fill()
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].alignment = get_centered_alignment()
    row += 1

    # Summary
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = get_subheader_font()
    ws[f'A{row}'].fill = get_subheader_fill()
    ws[f'A{row}'].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
    row += 1

    ws[f'A{row}'] = section_data.get('summary', '')
    ws[f'A{row}'].alignment = get_left_alignment()
    ws.merge_cells(f'A{row}:B{row}')
    ws.row_dimensions[row].height = 40
    row += 1

    # Score
    ws[f'A{row}'] = "Score"
    ws[f'B{row}'] = section_data.get('score', '—')
    ws[f'A{row}'].font = get_subheader_font()
    ws[f'B{row}'].font = get_body_font()
    ws[f'B{row}'].fill = PatternFill(start_color=score_to_color(section_data.get('score')),
                                      end_color=score_to_color(section_data.get('score')),
                                      fill_type="solid")
    row += 1

    # Strengths
    ws[f'A{row}'] = "Strengths"
    ws[f'A{row}'].font = get_subheader_font()
    ws[f'A{row}'].fill = get_accent_fill()
    ws[f'A{row}'].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
    row += 1

    for strength in section_data.get('strengths', []):
        ws[f'A{row}'] = "•"
        ws[f'B{row}'] = strength.get('statement', '') if isinstance(strength, dict) else str(strength)
        ws[f'B{row}'].alignment = get_left_alignment()
        ws[f'A{row}'].font = get_body_font()
        ws[f'B{row}'].font = get_body_font()
        row += 1

    row += 1

    # Weaknesses
    ws[f'A{row}'] = "Weaknesses"
    ws[f'A{row}'].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
    ws[f'A{row}'].fill = get_accent_fill()
    row += 1

    for weakness in section_data.get('weaknesses', []):
        ws[f'A{row}'] = "•"
        ws[f'B{row}'] = weakness.get('statement', '') if isinstance(weakness, dict) else str(weakness)
        ws[f'B{row}'].alignment = get_left_alignment()
        ws[f'A{row}'].font = get_body_font()
        ws[f'B{row}'].font = get_body_font()
        row += 1

    row += 1

    # Recommendations
    ws[f'A{row}'] = "Recommendations"
    ws[f'A{row}'].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
    ws[f'A{row}'].fill = get_accent_fill()
    row += 1

    for rec in section_data.get('recommendations', []):
        ws[f'A{row}'] = rec.get('priority', '').upper() if isinstance(rec, dict) else ""
        ws[f'B{row}'] = rec.get('action', '') if isinstance(rec, dict) else str(rec)
        ws[f'A{row}'].font = Font(name=FONT_FAMILY, size=9, bold=True, color=COLOR_DARK_CYAN)
        ws[f'B{row}'].alignment = get_left_alignment()
        ws[f'B{row}'].font = get_body_font()
        ws.row_dimensions[row].height = 30
        row += 1


def create_customer_journey_sheet(wb, journey_data):
    """Create the Customer Journey sheet with stage table."""
    ws = wb.create_sheet("→ Customer Journey")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30

    row = 1

    # Title
    ws[f'A{row}'] = "Customer Journey"
    ws[f'A{row}'].font = get_header_font()
    ws[f'A{row}'].fill = get_header_fill()
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].alignment = get_centered_alignment()
    row += 1

    # Summary
    ws[f'A{row}'] = journey_data.get('summary', '')
    ws[f'A{row}'].alignment = get_left_alignment()
    ws.merge_cells(f'A{row}:C{row}')
    ws.row_dimensions[row].height = 40
    row += 1

    # Score
    ws[f'A{row}'] = "Score"
    ws[f'B{row}'] = journey_data.get('score', '—')
    ws[f'B{row}'].fill = PatternFill(start_color=score_to_color(journey_data.get('score')),
                                      end_color=score_to_color(journey_data.get('score')),
                                      fill_type="solid")
    row += 2

    # Journey Stages Table
    ws[f'A{row}'] = "Stage"
    ws[f'B{row}'] = "Touchpoints"
    ws[f'C{row}'] = "Pain Points"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = get_header_font()
        ws[f'{col}{row}'].fill = get_header_fill()
        ws[f'{col}{row}'].alignment = get_centered_alignment()
    row += 1

    for stage in journey_data.get('journey_stages', []):
        ws[f'A{row}'] = stage.get('stage', '').title()
        touchpoints = "; ".join([tp.get('touchpoint', '') for tp in stage.get('touchpoints', [])])
        pain_points = "; ".join(stage.get('pain_points', []))
        ws[f'B{row}'] = touchpoints
        ws[f'C{row}'] = pain_points

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = get_body_font()
            ws[f'{col}{row}'].alignment = get_left_alignment()
            ws[f'{col}{row}'].border = get_thin_border()
        ws.row_dimensions[row].height = 30
        row += 1


def create_budget_sheet(wb, budget_data):
    """Create the Budget Analysis sheet with allocation table."""
    ws = wb.create_sheet("→ Budget")

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20

    row = 1

    # Title
    ws[f'A{row}'] = "Budget Analysis"
    ws[f'A{row}'].font = get_header_font()
    ws[f'A{row}'].fill = get_header_fill()
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].alignment = get_centered_alignment()
    row += 1

    # Summary
    ws[f'A{row}'] = budget_data.get('summary', '')
    ws[f'A{row}'].alignment = get_left_alignment()
    ws.merge_cells(f'A{row}:C{row}')
    ws.row_dimensions[row].height = 40
    row += 1

    # Total Spend
    ws[f'A{row}'] = "Total Marketing Spend"
    ws[f'B{row}'] = budget_data.get('total_marketing_spend', '—')
    ws[f'A{row}'].font = get_subheader_font()
    ws[f'B{row}'].font = get_body_font()
    row += 1

    ws[f'A{row}'] = "% of Revenue"
    ws[f'B{row}'] = budget_data.get('marketing_spend_as_pct_revenue', '—')
    ws[f'A{row}'].font = get_subheader_font()
    ws[f'B{row}'].font = get_body_font()
    row += 2

    # Channel Allocation Table
    ws[f'A{row}'] = "Channel"
    ws[f'B{row}'] = "% of Budget"
    ws[f'C{row}'] = "ROI Assessment"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = get_header_font()
        ws[f'{col}{row}'].fill = get_header_fill()
        ws[f'{col}{row}'].alignment = get_centered_alignment()
    row += 1

    for channel in budget_data.get('channel_allocation', []):
        ws[f'A{row}'] = channel.get('channel', '')
        ws[f'B{row}'] = f"{channel.get('pct_of_budget', 0)}%"
        ws[f'C{row}'] = channel.get('roi_assessment', '—').title()

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = get_body_font()
            ws[f'{col}{row}'].alignment = get_left_alignment()
            ws[f'{col}{row}'].border = get_thin_border()
        row += 1


def create_performance_sheet(wb, perf_data):
    """Create the Performance Metrics sheet with KPI table."""
    ws = wb.create_sheet("→ Performance")

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    row = 1

    # Title
    ws[f'A{row}'] = "Performance Metrics"
    ws[f'A{row}'].font = get_header_font()
    ws[f'A{row}'].fill = get_header_fill()
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].alignment = get_centered_alignment()
    row += 1

    # Summary
    ws[f'A{row}'] = perf_data.get('summary', '')
    ws[f'A{row}'].alignment = get_left_alignment()
    ws.merge_cells(f'A{row}:C{row}')
    ws.row_dimensions[row].height = 40
    row += 1

    # Attribution Model
    ws[f'A{row}'] = "Attribution Model"
    ws[f'B{row}'] = perf_data.get('attribution_model', '—').title()
    ws[f'A{row}'].font = get_subheader_font()
    ws[f'B{row}'].font = get_body_font()
    row += 2

    # KPIs Table
    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'C{row}'] = "Trend"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = get_header_font()
        ws[f'{col}{row}'].fill = get_header_fill()
        ws[f'{col}{row}'].alignment = get_centered_alignment()
    row += 1

    for kpi in perf_data.get('key_kpis', []):
        ws[f'A{row}'] = kpi.get('metric', '')
        ws[f'B{row}'] = kpi.get('value', '—')
        ws[f'C{row}'] = kpi.get('trend', '—').title()

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = get_body_font()
            ws[f'{col}{row}'].alignment = get_left_alignment()
            ws[f'{col}{row}'].border = get_thin_border()
        row += 1


def create_dynamics_sheet(wb, audit):
    """Create the Cross-Section Dynamics sheet."""
    ws = wb.create_sheet("Dynamics")

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    row = 1

    # Title
    ws[f'A{row}'] = "Cross-Section Dynamics"
    ws[f'A{row}'].font = get_header_font()
    ws[f'A{row}'].fill = get_header_fill()
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].alignment = get_centered_alignment()
    row += 2

    # Dynamics
    ws[f'A{row}'] = "Dynamic"
    ws[f'B{row}'] = "Insight"
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = get_header_font()
        ws[f'{col}{row}'].fill = get_header_fill()
        ws[f'{col}{row}'].alignment = get_centered_alignment()
    row += 1

    for dynamic in audit.get('cross_section_dynamics', []):
        ws[f'A{row}'] = dynamic.get('dynamic', '')
        ws[f'B{row}'] = dynamic.get('insight', '')

        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = get_body_font()
            ws[f'{col}{row}'].alignment = get_left_alignment()
            ws[f'{col}{row}'].border = get_thin_border()
        ws.row_dimensions[row].height = 40
        row += 1

    row += 2

    # Top Priorities
    ws[f'A{row}'] = "Top Priorities"
    ws[f'A{row}'].font = get_subheader_font()
    ws[f'A{row}'].fill = get_accent_fill()
    ws[f'A{row}'].font = Font(name=FONT_FAMILY, size=10, bold=True, color=COLOR_WHITE)
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    for priority in audit.get('top_priorities', []):
        ws[f'A{row}'] = f"#{priority.get('priority', '')}"
        ws[f'B{row}'] = priority.get('action', '')
        ws[f'A{row}'].font = Font(name=FONT_FAMILY, size=9, bold=True, color=COLOR_DARK_CYAN)
        ws[f'B{row}'].alignment = get_left_alignment()
        ws[f'B{row}'].font = get_body_font()
        ws.row_dimensions[row].height = 30
        row += 1


def build_excel(json_path, output_path):
    """
    Build complete Excel workbook from JSON audit.

    Args:
        json_path: Path to input JSON file
        output_path: Path to output Excel file
    """
    audit = load_audit_json(json_path)

    # Create workbook
    wb = Workbook()

    # Create Overview (first sheet)
    create_overview_sheet(wb, audit)

    # Create detail sheets for each section
    create_detail_sheet(wb, "→ Strategy", audit.get('strategic_assessment', {}))
    create_detail_sheet(wb, "→ Product", audit.get('product', {}))
    create_detail_sheet(wb, "→ Price", audit.get('price', {}))
    create_detail_sheet(wb, "→ Place", audit.get('place', {}))
    create_detail_sheet(wb, "→ Promotion", audit.get('promotion', {}))
    create_detail_sheet(wb, "→ People", audit.get('people', {}))
    create_detail_sheet(wb, "→ Process", audit.get('process', {}))
    create_detail_sheet(wb, "→ Physical Evidence", audit.get('physical_evidence', {}))
    create_detail_sheet(wb, "→ Brand Health", audit.get('brand_health', {}))

    # Create specialized sheets
    create_customer_journey_sheet(wb, audit.get('customer_journey', {}))
    create_budget_sheet(wb, audit.get('budget_analysis', {}))
    create_performance_sheet(wb, audit.get('performance_metrics', {}))
    create_dynamics_sheet(wb, audit)

    # Save
    wb.save(output_path)
    print(f"✓ Excel workbook created: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python build_excel.py <json_audit_file> [output_path]")
        sys.exit(1)

    json_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else json_file.replace('.json', '.xlsx')

    if not Path(json_file).exists():
        print(f"Error: {json_file} not found")
        sys.exit(1)

    build_excel(json_file, output_file)
