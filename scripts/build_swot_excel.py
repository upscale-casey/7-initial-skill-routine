#!/usr/bin/env python3
"""
SWOT Analysis Excel Workbook Generator

Converts a JSON SWOT analysis (following the schema in references/schema.md)
into a branded Upscale Excel workbook (.xlsx) with seven sheets:
1. Overview - 2x2 matrix summary, scores, executive summary, top priorities
2. S Strengths - all strength items with evidence and scoring
3. W Weaknesses - all weakness items with evidence and scoring
4. O Opportunities - all opportunity items with evidence and scoring
5. T Threats - all threat items with evidence and scoring
6. TOWS Matrix - SO, WO, ST, WT strategies in structured layout
7. Evidence Map - traceability table linking items to sources

Uses Upscale brand colors:
- Eerie Black: #191919 (headers, text)
- Lime Green: #34C52A (accents, scores)
- Dark Cyan: #429792 (sub-sections)
- Ivory: #FEFFEA (backgrounds)

Font: Lexend throughout
"""

import json
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# Upscale Brand Colors
EERIE_BLACK = "191919"
LIME_GREEN = "34C52A"
DARK_CYAN = "429792"
IVORY = "FEFFEA"

# Font specs
HEADER_FONT = Font(name="Lexend", size=12, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Lexend", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Lexend", size=10, color="191919")
LABEL_FONT = Font(name="Lexend", size=9, color="191919")

# Fill colors
HEADER_FILL = PatternFill(fgColor=EERIE_BLACK, patternType="solid")
SUBHEADER_FILL = PatternFill(fgColor=DARK_CYAN, patternType="solid")
ACCENT_FILL = PatternFill(fgColor=LIME_GREEN, patternType="solid")
BACKGROUND_FILL = PatternFill(fgColor=IVORY, patternType="solid")

# Borders
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Alignment
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")


def load_swot_json(filepath):
    """Load SWOT analysis JSON from file."""
    with open(filepath, 'r') as f:
        return json.load(f)


def create_workbook(swot_data, output_filepath):
    """Create Excel workbook from SWOT data."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets in order
    create_overview_sheet(wb, swot_data)
    create_strengths_sheet(wb, swot_data)
    create_weaknesses_sheet(wb, swot_data)
    create_opportunities_sheet(wb, swot_data)
    create_threats_sheet(wb, swot_data)
    create_tows_matrix_sheet(wb, swot_data)
    create_evidence_map_sheet(wb, swot_data)

    # Save workbook
    wb.save(output_filepath)
    print(f"Excel workbook created: {output_filepath}")


def create_overview_sheet(wb, swot_data):
    """Create Overview sheet with 2x2 matrix, scores, executive summary, and top priorities."""
    ws = wb.create_sheet("Overview", 0)
    ws.sheet_properties.tabColor = LIME_GREEN

    row = 1

    # Title
    ws['A1'] = f"SWOT Analysis: {swot_data['company_name']}"
    ws['A1'].font = Font(name="Lexend", size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = HEADER_FILL
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = CENTER_ALIGN
    row = 2

    # Analysis metadata
    ws[f'A{row}'] = f"Industry: {swot_data['industry']}"
    ws[f'A{row}'].font = LABEL_FONT
    row += 1
    ws[f'A{row}'] = f"Analysis Date: {swot_data['analysis_date']}"
    ws[f'A{row}'].font = LABEL_FONT
    row += 2

    # 2x2 SWOT Matrix Summary
    ws[f'A{row}'] = "SWOT Summary Matrix"
    ws[f'A{row}'].font = SUBHEADER_FONT
    ws[f'A{row}'].fill = SUBHEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].alignment = CENTER_ALIGN
    row += 1

    # Matrix headers
    ws[f'B{row}'] = "Internal Factors"
    ws[f'B{row}'].font = HEADER_FONT
    ws[f'B{row}'].fill = HEADER_FILL
    ws[f'B{row}'].alignment = CENTER_ALIGN
    ws.merge_cells(f'B{row}:C{row}')

    ws[f'D{row}'] = "External Factors"
    ws[f'D{row}'].font = HEADER_FONT
    ws[f'D{row}'].fill = HEADER_FILL
    ws[f'D{row}'].alignment = CENTER_ALIGN
    row += 1

    # S-W row header
    ws[f'A{row}'] = "Positive"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = ACCENT_FILL
    ws[f'A{row}'].alignment = CENTER_ALIGN
    row_positive = row

    # Strengths cell
    s_count = len(swot_data.get('strengths', []))
    ws[f'B{row}'] = f"Strengths\n({s_count})"
    ws[f'B{row}'].font = BODY_FONT
    ws[f'B{row}'].fill = BACKGROUND_FILL
    ws[f'B{row}'].alignment = CENTER_ALIGN
    ws[f'B{row}'].border = THIN_BORDER

    # Opportunities cell
    o_count = len(swot_data.get('opportunities', []))
    ws[f'D{row}'] = f"Opportunities\n({o_count})"
    ws[f'D{row}'].font = BODY_FONT
    ws[f'D{row}'].fill = BACKGROUND_FILL
    ws[f'D{row}'].alignment = CENTER_ALIGN
    ws[f'D{row}'].border = THIN_BORDER
    row += 1

    # W-T row
    ws[f'A{row}'] = "Negative"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = PatternFill(fgColor="D32F2F", patternType="solid")
    ws[f'A{row}'].alignment = CENTER_ALIGN

    # Weaknesses cell
    w_count = len(swot_data.get('weaknesses', []))
    ws[f'B{row}'] = f"Weaknesses\n({w_count})"
    ws[f'B{row}'].font = BODY_FONT
    ws[f'B{row}'].fill = BACKGROUND_FILL
    ws[f'B{row}'].alignment = CENTER_ALIGN
    ws[f'B{row}'].border = THIN_BORDER

    # Threats cell
    t_count = len(swot_data.get('threats', []))
    ws[f'D{row}'] = f"Threats\n({t_count})"
    ws[f'D{row}'].font = BODY_FONT
    ws[f'D{row}'].fill = BACKGROUND_FILL
    ws[f'D{row}'].alignment = CENTER_ALIGN
    ws[f'D{row}'].border = THIN_BORDER
    row += 2

    # Strategic Scores
    ws[f'A{row}'] = "Strategic Scores"
    ws[f'A{row}'].font = SUBHEADER_FONT
    ws[f'A{row}'].fill = SUBHEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].alignment = CENTER_ALIGN
    row += 1

    ws[f'A{row}'] = "Internal Score (S vs W):"
    ws[f'A{row}'].font = LABEL_FONT
    ws[f'B{row}'] = swot_data['internal_score']
    ws[f'B{row}'].font = Font(name="Lexend", size=11, bold=True, color="FFFFFF")
    ws[f'B{row}'].fill = ACCENT_FILL
    ws[f'B{row}'].alignment = CENTER_ALIGN
    row += 1

    ws[f'A{row}'] = "External Score (O vs T):"
    ws[f'A{row}'].font = LABEL_FONT
    ws[f'B{row}'] = swot_data['external_score']
    ws[f'B{row}'].font = Font(name="Lexend", size=11, bold=True, color="FFFFFF")
    ws[f'B{row}'].fill = ACCENT_FILL
    ws[f'B{row}'].alignment = CENTER_ALIGN
    row += 1

    ws[f'A{row}'] = "Overall Score:"
    ws[f'A{row}'].font = LABEL_FONT
    ws[f'B{row}'] = swot_data['overall_score']
    ws[f'B{row}'].font = Font(name="Lexend", size=11, bold=True, color="FFFFFF")
    ws[f'B{row}'].fill = ACCENT_FILL
    ws[f'B{row}'].alignment = CENTER_ALIGN
    row += 1

    ws[f'A{row}'] = "Strategic Position:"
    ws[f'A{row}'].font = LABEL_FONT
    ws[f'B{row}'] = swot_data['overall_strategic_position'].upper()
    ws[f'B{row}'].font = Font(name="Lexend", size=11, bold=True, color="FFFFFF")
    ws[f'B{row}'].fill = ACCENT_FILL
    ws[f'B{row}'].alignment = CENTER_ALIGN
    row += 2

    # Executive Summary
    ws[f'A{row}'] = "Executive Summary"
    ws[f'A{row}'].font = SUBHEADER_FONT
    ws[f'A{row}'].fill = SUBHEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].alignment = CENTER_ALIGN
    row += 1

    ws[f'A{row}'] = swot_data['executive_summary']
    ws[f'A{row}'].font = BODY_FONT
    ws[f'A{row}'].alignment = LEFT_ALIGN
    ws.merge_cells(f'A{row}:D{row+2}')
    ws.row_dimensions[row].height = 60
    row += 3

    # Top Priorities
    ws[f'A{row}'] = "Top Priorities"
    ws[f'A{row}'].font = SUBHEADER_FONT
    ws[f'A{row}'].fill = SUBHEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].alignment = CENTER_ALIGN
    row += 1

    # Priority headers
    headers = ["Priority", "Quadrant", "Statement", "Expected Impact"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    row += 1

    # Priority items
    for item in swot_data.get('top_priorities', [])[:5]:
        ws[f'A{row}'] = item['priority']
        ws[f'A{row}'].font = BODY_FONT
        ws[f'A{row}'].alignment = CENTER_ALIGN
        ws[f'A{row}'].fill = BACKGROUND_FILL

        ws[f'B{row}'] = item['quadrant_source'].upper()
        ws[f'B{row}'].font = BODY_FONT
        ws[f'B{row}'].alignment = CENTER_ALIGN
        ws[f'B{row}'].fill = BACKGROUND_FILL

        ws[f'C{row}'] = item['statement'][:80]  # Truncate for display
        ws[f'C{row}'].font = BODY_FONT
        ws[f'C{row}'].alignment = LEFT_ALIGN
        ws[f'C{row}'].fill = BACKGROUND_FILL
        ws[f'C{row}'].border = THIN_BORDER

        ws[f'D{row}'] = item['expected_impact'].upper()
        ws[f'D{row}'].font = BODY_FONT
        ws[f'D{row}'].alignment = CENTER_ALIGN
        ws[f'D{row}'].fill = BACKGROUND_FILL

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25


def create_quadrant_sheet(wb, swot_data, quadrant_key, quadrant_name, sheet_order):
    """Create a sheet for a SWOT quadrant (Strengths, Weaknesses, Opportunities, Threats)."""
    ws = wb.create_sheet(f"{sheet_order} {quadrant_name}", sheet_order)
    ws.sheet_properties.tabColor = LIME_GREEN

    items = swot_data.get(quadrant_key, [])

    row = 1

    # Title
    ws[f'A{row}'] = f"{quadrant_name} ({len(items)})"
    ws[f'A{row}'].font = Font(name="Lexend", size=13, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'].alignment = CENTER_ALIGN
    row += 1

    # Calculate average priority
    if items:
        avg_priority = sum(item['priority_score'] for item in items) / len(items)
        ws[f'A{row}'] = f"Count: {len(items)} | Avg Priority Score: {avg_priority:.1f}"
        ws[f'A{row}'].font = LABEL_FONT
        ws.merge_cells(f'A{row}:G{row}')
        row += 1

    row += 1

    # Table headers
    headers = ["Statement", "Evidence Source", "Evidence Detail", "Impact", "Likelihood", "Priority", "Confidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    row += 1

    # Data rows
    for item in items:
        ws[f'A{row}'] = item['statement']
        ws[f'A{row}'].font = BODY_FONT
        ws[f'A{row}'].alignment = LEFT_ALIGN
        ws[f'A{row}'].border = THIN_BORDER
        ws[f'A{row}'].fill = BACKGROUND_FILL
        ws.row_dimensions[row].height = 30

        ws[f'B{row}'] = item['evidence']['source_analysis']
        ws[f'B{row}'].font = BODY_FONT
        ws[f'B{row}'].alignment = CENTER_ALIGN
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'B{row}'].fill = BACKGROUND_FILL

        ws[f'C{row}'] = item['evidence']['detail'][:100]  # Truncate
        ws[f'C{row}'].font = LABEL_FONT
        ws[f'C{row}'].alignment = LEFT_ALIGN
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].fill = BACKGROUND_FILL

        ws[f'D{row}'] = item['impact_score']
        ws[f'D{row}'].font = Font(name="Lexend", size=10, bold=True, color="FFFFFF")
        ws[f'D{row}'].fill = ACCENT_FILL
        ws[f'D{row}'].alignment = CENTER_ALIGN
        ws[f'D{row}'].border = THIN_BORDER

        ws[f'E{row}'] = item['likelihood_score']
        ws[f'E{row}'].font = Font(name="Lexend", size=10, bold=True, color="FFFFFF")
        ws[f'E{row}'].fill = ACCENT_FILL
        ws[f'E{row}'].alignment = CENTER_ALIGN
        ws[f'E{row}'].border = THIN_BORDER

        ws[f'F{row}'] = item['priority_score']
        ws[f'F{row}'].font = Font(name="Lexend", size=10, bold=True, color="FFFFFF")
        ws[f'F{row}'].fill = SUBHEADER_FILL
        ws[f'F{row}'].alignment = CENTER_ALIGN
        ws[f'F{row}'].border = THIN_BORDER

        ws[f'G{row}'] = item['evidence']['confidence']
        ws[f'G{row}'].font = BODY_FONT
        ws[f'G{row}'].alignment = CENTER_ALIGN
        ws[f'G{row}'].border = THIN_BORDER
        ws[f'G{row}'].fill = BACKGROUND_FILL

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12


def create_strengths_sheet(wb, swot_data):
    """Create Strengths sheet."""
    create_quadrant_sheet(wb, swot_data, 'strengths', 'Strengths', 1)


def create_weaknesses_sheet(wb, swot_data):
    """Create Weaknesses sheet."""
    create_quadrant_sheet(wb, swot_data, 'weaknesses', 'Weaknesses', 2)


def create_opportunities_sheet(wb, swot_data):
    """Create Opportunities sheet."""
    create_quadrant_sheet(wb, swot_data, 'opportunities', 'Opportunities', 3)


def create_threats_sheet(wb, swot_data):
    """Create Threats sheet."""
    create_quadrant_sheet(wb, swot_data, 'threats', 'Threats', 4)


def create_tows_matrix_sheet(wb, swot_data):
    """Create TOWS Matrix sheet with SO, WO, ST, WT strategies."""
    ws = wb.create_sheet("TOWS Matrix", 5)
    ws.sheet_properties.tabColor = LIME_GREEN

    row = 1

    # Title
    ws[f'A{row}'] = "TOWS Matrix: Strategic Strategies"
    ws[f'A{row}'].font = Font(name="Lexend", size=13, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].alignment = CENTER_ALIGN
    row += 2

    tows_matrix = swot_data.get('tows_matrix', {})

    # Process each TOWS quadrant
    for quadrant_type, quadrant_label in [
        ('so_strategies', 'SO Strategies (Strengths + Opportunities)'),
        ('wo_strategies', 'WO Strategies (Weaknesses + Opportunities)'),
        ('st_strategies', 'ST Strategies (Strengths + Threats)'),
        ('wt_strategies', 'WT Strategies (Weaknesses + Threats)')
    ]:
        strategies = tows_matrix.get(quadrant_type, [])

        # Quadrant header
        ws[f'A{row}'] = quadrant_label
        ws[f'A{row}'].font = SUBHEADER_FONT
        ws[f'A{row}'].fill = SUBHEADER_FILL
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].alignment = CENTER_ALIGN
        row += 1

        if not strategies:
            ws[f'A{row}'] = "(No strategies defined)"
            ws[f'A{row}'].font = LABEL_FONT
            row += 1
        else:
            # Strategy headers
            headers = ["Strategy Name", "Description", "Priority", "Effort", "Timeframe", "Expected Impact"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
            row += 1

            # Strategy rows
            for strategy in strategies:
                ws[f'A{row}'] = strategy['strategy_name']
                ws[f'A{row}'].font = BODY_FONT
                ws[f'A{row}'].alignment = LEFT_ALIGN
                ws[f'A{row}'].border = THIN_BORDER
                ws[f'A{row}'].fill = BACKGROUND_FILL

                ws[f'B{row}'] = strategy['description'][:100]  # Truncate
                ws[f'B{row}'].font = LABEL_FONT
                ws[f'B{row}'].alignment = LEFT_ALIGN
                ws[f'B{row}'].border = THIN_BORDER
                ws[f'B{row}'].fill = BACKGROUND_FILL
                ws.row_dimensions[row].height = 30

                ws[f'C{row}'] = strategy['priority'].upper()
                ws[f'C{row}'].font = Font(name="Lexend", size=10, bold=True, color="FFFFFF")
                ws[f'C{row}'].fill = ACCENT_FILL
                ws[f'C{row}'].alignment = CENTER_ALIGN
                ws[f'C{row}'].border = THIN_BORDER

                ws[f'D{row}'] = strategy['effort'].upper()
                ws[f'D{row}'].font = BODY_FONT
                ws[f'D{row}'].alignment = CENTER_ALIGN
                ws[f'D{row}'].border = THIN_BORDER
                ws[f'D{row}'].fill = BACKGROUND_FILL

                ws[f'E{row}'] = strategy['timeframe'].replace('_', ' ').upper()
                ws[f'E{row}'].font = BODY_FONT
                ws[f'E{row}'].alignment = CENTER_ALIGN
                ws[f'E{row}'].border = THIN_BORDER
                ws[f'E{row}'].fill = BACKGROUND_FILL

                ws[f'F{row}'] = strategy['expected_impact'].upper()
                ws[f'F{row}'].font = BODY_FONT
                ws[f'F{row}'].alignment = CENTER_ALIGN
                ws[f'F{row}'].border = THIN_BORDER
                ws[f'F{row}'].fill = BACKGROUND_FILL

                row += 1

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16


def create_evidence_map_sheet(wb, swot_data):
    """Create Evidence Map sheet for traceability."""
    ws = wb.create_sheet("Evidence Map", 6)
    ws.sheet_properties.tabColor = LIME_GREEN

    row = 1

    # Title
    ws[f'A{row}'] = "Evidence Map: Item Traceability"
    ws[f'A{row}'].font = Font(name="Lexend", size=13, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].alignment = CENTER_ALIGN
    row += 2

    # Table headers
    headers = ["Statement", "Quadrant", "Source Analysis", "Source Element", "Evidence Detail", "Confidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    row += 1

    # Collect all items from all quadrants
    all_items = []
    for quadrant_key, quadrant_name in [('strengths', 'Strength'), ('weaknesses', 'Weakness'),
                                          ('opportunities', 'Opportunity'), ('threats', 'Threat')]:
        for item in swot_data.get(quadrant_key, []):
            all_items.append((quadrant_name, item))

    # Add rows
    for quadrant_name, item in all_items:
        ws[f'A{row}'] = item['statement'][:80]  # Truncate
        ws[f'A{row}'].font = LABEL_FONT
        ws[f'A{row}'].alignment = LEFT_ALIGN
        ws[f'A{row}'].border = THIN_BORDER
        ws[f'A{row}'].fill = BACKGROUND_FILL
        ws.row_dimensions[row].height = 25

        ws[f'B{row}'] = quadrant_name
        ws[f'B{row}'].font = BODY_FONT
        ws[f'B{row}'].alignment = CENTER_ALIGN
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'B{row}'].fill = BACKGROUND_FILL

        ws[f'C{row}'] = item['evidence']['source_analysis']
        ws[f'C{row}'].font = BODY_FONT
        ws[f'C{row}'].alignment = CENTER_ALIGN
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].fill = BACKGROUND_FILL

        ws[f'D{row}'] = item['evidence']['source_element']
        ws[f'D{row}'].font = LABEL_FONT
        ws[f'D{row}'].alignment = LEFT_ALIGN
        ws[f'D{row}'].border = THIN_BORDER
        ws[f'D{row}'].fill = BACKGROUND_FILL

        ws[f'E{row}'] = item['evidence']['detail'][:80]  # Truncate
        ws[f'E{row}'].font = LABEL_FONT
        ws[f'E{row}'].alignment = LEFT_ALIGN
        ws[f'E{row}'].border = THIN_BORDER
        ws[f'E{row}'].fill = BACKGROUND_FILL

        ws[f'F{row}'] = item['evidence']['confidence'].upper()
        ws[f'F{row}'].font = BODY_FONT
        ws[f'F{row}'].alignment = CENTER_ALIGN
        ws[f'F{row}'].border = THIN_BORDER
        ws[f'F{row}'].fill = BACKGROUND_FILL

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage: python build_excel.py <swot_json_filepath> [output_filepath]")
        print("\nExample:")
        print("  python build_excel.py swot_analysis.json swot_output.xlsx")
        sys.exit(1)

    json_filepath = sys.argv[1]
    output_filepath = sys.argv[2] if len(sys.argv) > 2 else "swot_analysis.xlsx"

    try:
        swot_data = load_swot_json(json_filepath)
        create_workbook(swot_data, output_filepath)
        print(f"Success! SWOT workbook saved to: {output_filepath}")
    except FileNotFoundError:
        print(f"Error: JSON file not found: {json_filepath}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON in {json_filepath}: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
