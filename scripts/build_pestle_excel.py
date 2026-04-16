#!/usr/bin/env python3
"""
PESTLE Analysis Excel Builder

Generates a branded Upscale Excel workbook (.xlsx) from PESTLE JSON analysis output.
Follows the same pattern as the 4ps-excel build script.

Usage:
    python build_excel.py <input_json> <output_xlsx>

Example:
    python build_excel.py pestle_analysis.json pestle_report.xlsx
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        DEFAULT_FONT
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Upscale Brand Palette
COLORS = {
    "eerie_black": "191919",
    "lime_green": "34C52A",
    "dark_cyan": "429792",
    "ivory": "FEFFEA",
    "light_gray": "F5F5F5",
    "white": "FFFFFF",
}

# Risk Zone Colors
RISK_COLORS = {
    "low": "C6EFCE",  # Green tint
    "medium": "B4C7E7",  # Cyan tint
    "high": "F5A623",  # Orange
    "critical": "D5B5DB",  # Purple tint
}

# Helper Functions
def hex_to_rgb(hex_color):
    """Convert hex color to RGB tuple."""
    hex_color = hex_color.lstrip("#")
    return tuple(int(hex_color[i : i + 2], 16) for i in (0, 2, 4))


def set_cell_style(cell, font_color=None, fill_color=None, bold=False, size=11, alignment="left"):
    """Apply consistent styling to a cell."""
    if font_color:
        cell.font = Font(name="Lexend", color=font_color, bold=bold, size=size)
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if alignment == "center":
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    elif alignment == "left":
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    elif alignment == "right":
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)


def add_header_row(ws, row, values, font_color=COLORS["white"], fill_color=COLORS["eerie_black"]):
    """Add a styled header row."""
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        set_cell_style(cell, font_color=font_color, fill_color=fill_color, bold=True, size=12, alignment="center")


def add_score_bar(ws, row, col, score, max_score=10):
    """Add a visual score bar (filled cells)."""
    for i in range(1, max_score + 1):
        cell = ws.cell(row=row, column=col + i)
        if i <= score:
            cell.fill = PatternFill(start_color=COLORS["lime_green"], end_color=COLORS["lime_green"], fill_type="solid")
        else:
            cell.fill = PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(size=8)


# Sheet Builders
def build_overview_sheet(ws, data):
    """Build the Overview sheet."""
    ws.title = "Overview"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 15

    # Title
    title_cell = ws.cell(row=1, column=1, value="PESTLE ANALYSIS OVERVIEW")
    set_cell_style(title_cell, font_color=COLORS["white"], fill_color=COLORS["eerie_black"], bold=True, size=15, alignment="center")
    ws.merge_cells("A1:C1")

    # Company & Industry
    row = 3
    ws.cell(row=row, column=1, value="Company").font = Font(bold=True, name="Lexend")
    ws.cell(row=row, column=2, value=data.get("company_name", ""))

    row += 1
    ws.cell(row=row, column=1, value="Industry").font = Font(bold=True, name="Lexend")
    ws.cell(row=row, column=2, value=data.get("industry", ""))

    row += 1
    ws.cell(row=row, column=1, value="Market Scope").font = Font(bold=True, name="Lexend")
    ws.cell(row=row, column=2, value=data.get("market_scope", ""))

    row += 1
    ws.cell(row=row, column=1, value="Analysis Date").font = Font(bold=True, name="Lexend")
    ws.cell(row=row, column=2, value=data.get("analysis_date", ""))

    # Scores
    row += 2
    add_header_row(ws, row, ["Metric", "Score", "Visual"])
    row += 1

    ws.cell(row=row, column=1, value="Overall Risk Score").font = Font(bold=True, name="Lexend")
    ws.cell(row=row, column=2, value=data.get("overall_risk_score", 0))
    add_score_bar(ws, row, 2, data.get("overall_risk_score", 0), 10)

    row += 1
    ws.cell(row=row, column=1, value="Overall Opportunity Score").font = Font(bold=True, name="Lexend")
    ws.cell(row=row, column=2, value=data.get("overall_opportunity_score", 0))
    add_score_bar(ws, row, 2, data.get("overall_opportunity_score", 0), 10)

    # Overall Assessment
    row += 2
    ws.cell(row=row, column=1, value="Overall Assessment").font = Font(bold=True, name="Lexend")
    assessment_cell = ws.cell(row=row, column=2, value=data.get("overall_assessment", "").replace("_", " ").title())
    set_cell_style(assessment_cell, fill_color=COLORS["lime_green"], font_color=COLORS["eerie_black"], bold=True)

    # Executive Summary
    row += 2
    summary_cell = ws.cell(row=row, column=1, value="Executive Summary")
    set_cell_style(summary_cell, font_color=COLORS["white"], fill_color=COLORS["dark_cyan"], bold=True, size=12)
    ws.merge_cells(f"A{row}:C{row}")

    row += 1
    summary_text_cell = ws.cell(row=row, column=1, value=data.get("executive_summary", ""))
    summary_text_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row}:C{row + 2}")
    ws.row_dimensions[row].height = 60

    # Risk Matrix Table (abbreviated)
    row += 4
    add_header_row(ws, row, ["Factor", "Risk Zone", "Priority"])
    row += 1

    for factor in data.get("risk_matrix", [])[:10]:  # Top 10
        ws.cell(row=row, column=1, value=factor.get("factor_name", ""))
        risk_zone_cell = ws.cell(row=row, column=2, value=factor.get("risk_zone", "").upper())
        set_cell_style(risk_zone_cell, fill_color=RISK_COLORS.get(factor.get("risk_zone", "low"), COLORS["light_gray"]))
        ws.cell(row=row, column=3, value=factor.get("priority_score", 0))
        row += 1

    # Top Priorities
    row += 2
    priorities_cell = ws.cell(row=row, column=1, value="Top Priorities")
    set_cell_style(priorities_cell, font_color=COLORS["white"], fill_color=COLORS["dark_cyan"], bold=True, size=12)
    ws.merge_cells(f"A{row}:C{row}")

    row += 1
    for idx, priority in enumerate(data.get("top_priorities", [])[:5], 1):
        priority_text = f"{idx}. [{priority.get('pestle_category', '').upper()}] {priority.get('factor_name', '')} ({priority.get('expected_impact', '').title()})"
        ws.cell(row=row, column=1, value=priority_text)
        ws.merge_cells(f"A{row}:C{row}")
        row += 1


def build_pestle_factor_sheet(ws, factor_key, factor_data):
    """Build a single PESTLE factor sheet."""
    factor_titles = {
        "political": "P - Political",
        "economic": "E - Economic",
        "social": "S - Social",
        "technological": "T - Technological",
        "legal": "L - Legal",
        "environmental": "E - Environmental",
    }

    ws.title = factor_titles.get(factor_key, factor_key)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15

    # Title
    title_cell = ws.cell(row=1, column=1, value=f"{factor_titles.get(factor_key, factor_key).upper()}")
    set_cell_style(title_cell, font_color=COLORS["white"], fill_color=COLORS["eerie_black"], bold=True, size=14, alignment="center")
    ws.merge_cells("A1:F1")

    # Factor Score Bar
    row = 2
    score = factor_data.get("score", 5)
    ws.cell(row=row, column=1, value=f"Factor Score: {score}/10").font = Font(bold=True, name="Lexend")
    add_score_bar(ws, row, 2, score, 10)

    # Summary
    row += 2
    summary_cell = ws.cell(row=row, column=1, value="Summary")
    set_cell_style(summary_cell, font_color=COLORS["white"], fill_color=COLORS["dark_cyan"], bold=True)
    ws.merge_cells(f"A{row}:F{row}")

    row += 1
    summary_text = ws.cell(row=row, column=1, value=factor_data.get("summary", ""))
    summary_text.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"A{row}:F{row + 1}")
    ws.row_dimensions[row].height = 40

    # Factors Table
    row += 3
    add_header_row(ws, row, ["Trend/Force", "Description", "Impact", "Likelihood", "Priority", "Risk Zone"])
    row += 1

    for factor in factor_data.get("factors", []):
        ws.cell(row=row, column=1, value=factor.get("factor_name", ""))
        ws.cell(row=row, column=2, value=factor.get("description", ""))
        ws.cell(row=row, column=3, value=factor.get("impact_score", ""))
        ws.cell(row=row, column=4, value=factor.get("likelihood_score", ""))
        ws.cell(row=row, column=5, value=factor.get("priority_score", ""))
        risk_zone_cell = ws.cell(row=row, column=6, value=factor.get("risk_zone", "").upper())
        set_cell_style(risk_zone_cell, fill_color=RISK_COLORS.get(factor.get("risk_zone", "low"), COLORS["light_gray"]))
        for col in range(1, 7):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    # Key Risks
    row += 2
    risks_cell = ws.cell(row=row, column=1, value="Key Risks")
    set_cell_style(risks_cell, font_color=COLORS["white"], fill_color=COLORS["dark_cyan"], bold=True)
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    for risk in factor_data.get("key_risks", []):
        ws.cell(row=row, column=1, value=f"• {risk}")
        ws.merge_cells(f"A{row}:F{row}")
        ws.row_dimensions[row].height = 20
        row += 1

    # Key Opportunities
    row += 1
    opp_cell = ws.cell(row=row, column=1, value="Key Opportunities")
    opp_cell.font = Font(name="Lexend", color=COLORS["eerie_black"], bold=True, size=11)
    opp_cell.fill = PatternFill(start_color=COLORS["lime_green"], end_color=COLORS["lime_green"], fill_type="solid")
    opp_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    for opp in factor_data.get("key_opportunities", []):
        ws.cell(row=row, column=1, value=f"• {opp}")
        ws.merge_cells(f"A{row}:F{row}")
        ws.row_dimensions[row].height = 20
        row += 1

    # Recommendations
    row += 1
    rec_cell = ws.cell(row=row, column=1, value="Recommendations")
    set_cell_style(rec_cell, font_color=COLORS["white"], fill_color=COLORS["dark_cyan"], bold=True)
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    for rec in factor_data.get("recommendations", []):
        action = rec.get("action", "")
        priority = rec.get("priority", "medium")
        ws.cell(row=row, column=1, value=f"• {action}")
        priority_cell = ws.cell(row=row, column=6, value=priority.upper())
        set_cell_style(priority_cell, bold=True, fill_color=COLORS["light_gray"])
        ws.merge_cells(f"A{row}:E{row}")
        ws.row_dimensions[row].height = 20
        row += 1


def build_risk_matrix_sheet(ws, data):
    """Build the Risk Matrix sheet with sortable table."""
    ws.title = "Risk Matrix"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12

    # Title
    title_cell = ws.cell(row=1, column=1, value="PESTLE RISK MATRIX")
    set_cell_style(title_cell, font_color=COLORS["white"], fill_color=COLORS["eerie_black"], bold=True, size=14, alignment="center")
    ws.merge_cells("A1:F1")

    # Headers
    row = 3
    add_header_row(ws, row, ["Factor", "Category", "Impact", "Likelihood", "Priority", "Risk Zone"])

    # Data rows
    row += 1
    start_row = row
    for factor in data.get("risk_matrix", []):
        ws.cell(row=row, column=1, value=factor.get("factor_name", ""))
        ws.cell(row=row, column=2, value=factor.get("category", "").upper())
        ws.cell(row=row, column=3, value=factor.get("impact_score", ""))
        ws.cell(row=row, column=4, value=factor.get("likelihood_score", ""))
        ws.cell(row=row, column=5, value=factor.get("priority_score", ""))
        risk_zone_cell = ws.cell(row=row, column=6, value=factor.get("risk_zone", "").upper())
        set_cell_style(risk_zone_cell, fill_color=RISK_COLORS.get(factor.get("risk_zone", "low"), COLORS["light_gray"]))
        row += 1

    # Create table for sorting
    try:
        tab = Table(displayName="RiskMatrix", ref=f"A3:F{row - 1}")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except:
        pass  # Skip table if openpyxl version doesn't support it


def build_swot_feed_sheet(ws, data):
    """Build the SWOT Feed sheet."""
    ws.title = "SWOT Feed"
    ws.column_dimensions["A"].width = 60

    # Title
    title_cell = ws.cell(row=1, column=1, value="SWOT FEED FROM PESTLE")
    set_cell_style(title_cell, font_color=COLORS["white"], fill_color=COLORS["eerie_black"], bold=True, size=14, alignment="center")
    ws.merge_cells("A1:A1")

    # Opportunities
    row = 3
    opp_cell = ws.cell(row=row, column=1, value="OPPORTUNITIES")
    opp_cell.font = Font(name="Lexend", color=COLORS["eerie_black"], bold=True, size=12)
    opp_cell.fill = PatternFill(start_color=COLORS["lime_green"], end_color=COLORS["lime_green"], fill_type="solid")
    opp_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    row += 1

    for opp in data.get("swot_feed", {}).get("opportunities", []):
        ws.cell(row=row, column=1, value=f"• {opp}")
        ws.row_dimensions[row].height = 25
        row += 1

    # Threats
    row += 2
    threat_cell = ws.cell(row=row, column=1, value="THREATS")
    threat_cell.font = Font(name="Lexend", color=COLORS["eerie_black"], bold=True, size=12)
    threat_cell.fill = PatternFill(start_color=RISK_COLORS["critical"], end_color=RISK_COLORS["critical"], fill_type="solid")
    threat_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    row += 1

    for threat in data.get("swot_feed", {}).get("threats", []):
        ws.cell(row=row, column=1, value=f"• {threat}")
        ws.row_dimensions[row].height = 25
        row += 1


def build_cross_factor_sheet(ws, data):
    """Build the Cross-Factor Dynamics sheet."""
    ws.title = "Cross-Factor Dynamics"
    ws.column_dimensions["A"].width = 80

    # Title
    title_cell = ws.cell(row=1, column=1, value="CROSS-FACTOR DYNAMICS")
    set_cell_style(title_cell, font_color=COLORS["white"], fill_color=COLORS["eerie_black"], bold=True, size=14, alignment="center")
    ws.merge_cells("A1:A1")

    # Introduction
    row = 3
    intro_cell = ws.cell(row=row, column=1, value="Interactions & Dependencies Between PESTLE Factors")
    set_cell_style(intro_cell, font_color=COLORS["white"], fill_color=COLORS["dark_cyan"], bold=True)
    ws.merge_cells("A3:A3")

    # Dynamics
    row = 5
    for idx, dynamic in enumerate(data.get("cross_factor_dynamics", []), 1):
        ws.cell(row=row, column=1, value=f"{idx}. {dynamic}")
        ws.row_dimensions[row].height = 30
        row += 2


def build_workbook(input_json, output_xlsx):
    """Main function: build the complete Excel workbook."""
    # Load JSON
    with open(input_json, "r") as f:
        data = json.load(f)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Build all sheets
    ws = wb.create_sheet()
    build_overview_sheet(ws, data)

    factor_keys = ["political", "economic", "social", "technological", "legal", "environmental"]
    for key in factor_keys:
        if key in data:
            ws = wb.create_sheet()
            build_pestle_factor_sheet(ws, key, data[key])

    ws = wb.create_sheet()
    build_risk_matrix_sheet(ws, data)

    ws = wb.create_sheet()
    build_swot_feed_sheet(ws, data)

    ws = wb.create_sheet()
    build_cross_factor_sheet(ws, data)

    # Save
    wb.save(output_xlsx)
    print(f"Excel workbook created: {output_xlsx}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python build_excel.py <input_json> <output_xlsx>")
        sys.exit(1)

    input_json = sys.argv[1]
    output_xlsx = sys.argv[2]

    if not Path(input_json).exists():
        print(f"Error: Input file not found: {input_json}")
        sys.exit(1)

    try:
        build_workbook(input_json, output_xlsx)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
